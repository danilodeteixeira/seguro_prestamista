#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import http.server
import socketserver
import json
import urllib.parse
import math
import csv
import os
from pathlib import Path
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment
import multiprocessing as mp
from functools import lru_cache
from concurrent.futures import ProcessPoolExecutor, as_completed
import time
import socket

# Configuração do servidor
PORT_INICIAL = 8001

# Cache global para tábuas de comutação (otimização de performance)
TABUAS_CACHE = {}

def obter_tabua_cached(taxa_juros, tabua_nome):
    """Obtém tábua do cache ou cria nova se não existir"""
    cache_key = f"{taxa_juros}_{tabua_nome}"
    
    if cache_key not in TABUAS_CACHE:
        print(f"Carregando tábua {tabua_nome} no cache...")
        TABUAS_CACHE[cache_key] = TabuladeComutacao(taxa_juros, tabua_nome)
        print(f"Tábua {tabua_nome} carregada no cache!")
    
    return TABUAS_CACHE[cache_key]

class TabuladeComutacao:
    def __init__(self, taxa_juros, tabua_selecionada=None):
        self.taxa_juros = taxa_juros
        self.v = 1 / (1 + taxa_juros)
        self.tabuas_disponiveis = {}
        self.tabua_padrao = None
        self.carregar_todas_tabuas()
        self.tabua_selecionada = tabua_selecionada if tabua_selecionada else self.tabua_padrao
        self.dados = self.obter_dados_tabua()
        self.calcular_tabua_comutacao()
    
    def carregar_todas_tabuas(self):
        try:
            # Carregar dados das tábuas internas
            self.carregar_tabuas_internas()
        except Exception as e:
            print(f"Erro ao carregar tábuas: {e}. Usando tábua padrão.")
            self.carregar_tabua_padrao()
    
    def carregar_tabuas_internas(self):
        """Carrega as tábuas de mortalidade dos dados internos"""
        try:
            # Carregar dados do arquivo tabuas_mortalidade.js
            self.carregar_tabuas_js()
        except Exception as e:
            print(f"Erro ao carregar tábuas do JS: {e}. Usando tábua padrão.")
            self.carregar_tabua_padrao()
    
    def carregar_tabuas_js(self):
        """Carrega as tábuas de mortalidade do arquivo tabuas_mortalidade.js"""
        try:
            import json
            import re
            
            # Ler o arquivo tabuas_mortalidade.js
            with open('tabuas_mortalidade.js', 'r', encoding='utf-8') as f:
                content = f.read()
            
            print("Arquivo tabuas_mortalidade.js lido com sucesso")
            print(f"Tamanho do arquivo: {len(content)} caracteres")
            
            # Extrair o objeto TABUAS_MORTALIDADE usando regex mais robusto
            pattern = r'const TABUAS_MORTALIDADE\s*=\s*({.*?});'
            match = re.search(pattern, content, re.DOTALL)
            
            if match:
                print("Padrão TABUAS_MORTALIDADE encontrado no arquivo")
                js_data = match.group(1)
                print(f"Tamanho dos dados extraídos: {len(js_data)} caracteres")
                
                # Converter JavaScript para JSON de forma mais robusta
                # Primeiro, substituir chaves não quotadas por quotadas
                js_data = re.sub(r'(\w+):', r'"\1":', js_data)
                
                # Substituir aspas simples por aspas duplas
                js_data = js_data.replace("'", '"')
                
                # Converter para Python dict
                tabuas_data = json.loads(js_data)
                print(f"JSON parseado com sucesso. {len(tabuas_data)} tábuas encontradas")
                
                # Processar dados para o formato esperado
                for tabua_nome, dados in tabuas_data.items():
                    self.tabuas_disponiveis[tabua_nome] = dados
                    if self.tabua_padrao is None:
                        self.tabua_padrao = tabua_nome
                
                print(f"Tábuas carregadas: {list(self.tabuas_disponiveis.keys())}")
                
                # Testar acesso a uma tábua específica
                if 'AT-83' in self.tabuas_disponiveis:
                    at83_data = self.tabuas_disponiveis['AT-83']
                    if 'masculino' in at83_data and '30' in at83_data['masculino']:
                        qx_30_m = at83_data['masculino']['30']
                        print(f"Teste AT-83 - qx para idade 30, sexo M: {qx_30_m}")
                    else:
                        print("AT-83 encontrada mas estrutura de dados incorreta")
                else:
                    print("AT-83 não encontrada nas tábuas carregadas")
                    
            else:
                print("Padrão TABUAS_MORTALIDADE NÃO encontrado no arquivo")
                print("Primeiras 500 caracteres do arquivo:")
                print(content[:500])
                raise Exception("Não foi possível extrair TABUAS_MORTALIDADE do arquivo JS")
                
        except Exception as e:
            print(f"Erro ao processar tabuas_mortalidade.js: {e}")
            import traceback
            traceback.print_exc()
            raise
    
    def carregar_tabua_padrao(self):
        # Cria uma tábua padrão genérica
        tabua_padrao_nome = "Tábua Padrão"
        self.tabuas_disponiveis[tabua_padrao_nome] = {
            "masculino": {},
            "feminino": {}
        }
        dados_padrao = self.carregar_tabua_mortalidade()
        self.tabuas_disponiveis[tabua_padrao_nome] = dados_padrao
        self.tabua_padrao = tabua_padrao_nome
    
    def obter_dados_tabua(self):
        if self.tabua_selecionada in self.tabuas_disponiveis:
            return self.tabuas_disponiveis[self.tabua_selecionada]
        else:
            return self.carregar_tabua_mortalidade()
    
    def carregar_tabua_mortalidade(self):
        dados = {
            'masculino': {},
            'feminino': {}
        }
        # Dados da tábua BR-EMS 2021 (simplificada)
        tabua_data = [
            (0, 0.000371, 0.000355), (1, 0.000242, 0.000226), (2, 0.000213, 0.000195),
            (3, 0.000199, 0.000180), (4, 0.000192, 0.000171), (5, 0.000188, 0.000165),
            (6, 0.000186, 0.000162), (7, 0.000185, 0.000160), (8, 0.000185, 0.000159),
            (9, 0.000186, 0.000159), (10, 0.000188, 0.000160), (11, 0.000191, 0.000162),
            (12, 0.000195, 0.000165), (13, 0.000200, 0.000169), (14, 0.000206, 0.000174),
            (15, 0.000213, 0.000180), (16, 0.000221, 0.000187), (17, 0.000230, 0.000195),
            (18, 0.000240, 0.000204), (19, 0.000251, 0.000214), (20, 0.000263, 0.000225),
            (21, 0.000276, 0.000237), (22, 0.000290, 0.000250), (23, 0.000305, 0.000264),
            (24, 0.000321, 0.000279), (25, 0.000338, 0.000295), (26, 0.000356, 0.000312),
            (27, 0.000375, 0.000330), (28, 0.000395, 0.000349), (29, 0.000416, 0.000369),
            (30, 0.000438, 0.000390), (31, 0.000461, 0.000412), (32, 0.000485, 0.000435),
            (33, 0.000510, 0.000459), (34, 0.000536, 0.000484), (35, 0.000563, 0.000510),
            (36, 0.000591, 0.000537), (37, 0.000620, 0.000565), (38, 0.000650, 0.000594),
            (39, 0.000681, 0.000624), (40, 0.000713, 0.000655), (41, 0.000746, 0.000687),
            (42, 0.000780, 0.000720), (43, 0.000815, 0.000754), (44, 0.000851, 0.000789),
            (45, 0.000888, 0.000825), (46, 0.000926, 0.000862), (47, 0.000965, 0.000900),
            (48, 0.001005, 0.000939), (49, 0.001046, 0.000979), (50, 0.001088, 0.001020),
            (51, 0.001131, 0.001062), (52, 0.001175, 0.001105), (53, 0.001220, 0.001149),
            (54, 0.001266, 0.001194), (55, 0.001313, 0.001240), (56, 0.001361, 0.001287),
            (57, 0.001410, 0.001335), (58, 0.001460, 0.001384), (59, 0.001511, 0.001434),
            (60, 0.001563, 0.001485), (61, 0.001616, 0.001537), (62, 0.001670, 0.001590),
            (63, 0.001725, 0.001644), (64, 0.001781, 0.001699), (65, 0.001838, 0.001755),
            (66, 0.001896, 0.001812), (67, 0.001955, 0.001870), (68, 0.002015, 0.001989),
            (69, 0.002076, 0.001989), (70, 0.002138, 0.002050), (71, 0.002201, 0.002112),
            (72, 0.002265, 0.002175), (73, 0.002330, 0.002239), (74, 0.002396, 0.002304),
            (75, 0.002463, 0.002370), (76, 0.002531, 0.002437), (77, 0.002600, 0.002505),
            (78, 0.002670, 0.002574), (79, 0.002741, 0.002644), (80, 0.002813, 0.002715),
            (81, 0.002886, 0.002787), (82, 0.002960, 0.002860), (83, 0.003035, 0.002934),
            (84, 0.003111, 0.003009), (85, 0.003188, 0.003085), (86, 0.003266, 0.003162),
            (87, 0.003345, 0.003240), (88, 0.003425, 0.003319), (89, 0.003506, 0.003399),
            (90, 0.003588, 0.003480), (91, 0.003671, 0.003562), (92, 0.003755, 0.003645),
            (93, 0.003840, 0.003729), (94, 0.003926, 0.003814), (95, 0.004013, 0.003900),
            (96, 0.004101, 0.003987), (97, 0.004190, 0.004075), (98, 0.004280, 0.004164),
            (99, 0.004371, 0.004254), (100, 0.004463, 0.004345), (101, 0.004556, 0.004437),
            (102, 0.004650, 0.004530), (103, 0.004745, 0.004624), (104, 0.004841, 0.004719),
            (105, 0.004938, 0.004815), (106, 0.005036, 0.004912), (107, 0.005135, 0.005010),
            (108, 0.005235, 0.005109), (109, 0.005336, 0.005209), (110, 0.005438, 0.005310),
            (111, 0.005541, 0.005412), (112, 0.005645, 0.005515), (113, 0.005750, 0.005619),
            (114, 0.005856, 0.005724), (115, 0.005963, 0.005830), (116, 0.006071, 0.005937),
            (117, 0.006180, 0.006045), (118, 0.006290, 0.006154), (119, 0.006401, 0.006264),
            (120, 0.006513, 0.006375), (121, 0.006626, 0.006487), (122, 0.006740, 0.006600),
            (123, 0.006855, 0.006714), (124, 0.006971, 0.006829), (125, 1.000000, 1.000000)
        ]
        for idade, qx_masc, qx_fem in tabua_data:
            dados['masculino'][idade] = qx_masc
            dados['feminino'][idade] = qx_fem
        return dados
    
    def obter_qx(self, idade, sexo):
        # Converter idade para string para acessar os dados
        idade_str = str(idade)
        
        if sexo == 'M':
            return self.dados['masculino'].get(idade_str, 1.0)
        else:
            return self.dados['feminino'].get(idade_str, 1.0)
    
    def calcular_tabua_comutacao(self):
        self.l_x = {0: 100000}  # Radix
        self.d_x = {}
        self.D_x = {}
        self.C_x = {}
        self.N_x = {}
        self.M_x = {}
        self.v_x = {}
        
        sexo = 'M'  # Default para cálculo inicial, será ajustado na chamada
        
        # Calcular l_x e d_x
        for idade in range(0, 126):
            if idade > 0:
                qx_anterior = self.obter_qx(idade - 1, sexo)
                self.l_x[idade] = self.l_x[idade - 1] * (1 - qx_anterior)
            
            qx = self.obter_qx(idade, sexo)
            self.d_x[idade] = self.l_x[idade] * qx
        
        # Calcular D_x, C_x e v_x
        for idade in range(0, 126):
            self.D_x[idade] = self.l_x[idade] * (self.v ** idade)
            self.C_x[idade] = self.d_x[idade] * (self.v ** (idade + 1))
            self.v_x[idade] = self.v ** idade
        
        # Calcular N_x e M_x
        for idade in range(0, 126):
            self.N_x[idade] = sum(self.D_x.get(x, 0) for x in range(idade, 126))
            self.M_x[idade] = sum(self.C_x.get(x, 0) for x in range(idade, 126))

def calcular_seguro_anual(tabua_obj, idade, sexo, periodo):
    tabua_obj.dados = tabua_obj.tabuas_disponiveis[tabua_obj.tabua_selecionada]
    tabua_obj.calcular_tabua_comutacao()  # Recalcular com o sexo correto
    seguro_anual = (tabua_obj.M_x[idade] - tabua_obj.M_x[idade + periodo]) / tabua_obj.D_x[idade]
    return seguro_anual

def calcular_seguro_fracionado_total(taxa_juros, fracionamento, seguro_anual):
    taxa_fracionada = fracionamento * ((1 + taxa_juros)**(1/fracionamento) - 1)
    fator_ajuste = taxa_juros / taxa_fracionada
    seguro_fracionado_total = fator_ajuste * seguro_anual
    return seguro_fracionado_total, fator_ajuste, taxa_fracionada

def calcular_premio_mensal(tabua_obj, idade, periodo, valor_fracionado_total, taxa_juros, fracionamento, soma_segurada):
    N_x = tabua_obj.N_x[idade]
    N_x_n = tabua_obj.N_x[idade + periodo]
    D_x = tabua_obj.D_x[idade]
    D_x_n = tabua_obj.D_x[idade + periodo]
    
    anuidade_ajustada = ((N_x - N_x_n) / D_x + (11/24 * (1 - D_x_n / D_x))) * fracionamento
    valor_mensal = valor_fracionado_total / anuidade_ajustada if anuidade_ajustada != 0 else 0
    
    percentual_mensal = calcular_percentual_mensal(valor_mensal, taxa_juros, fracionamento, periodo, soma_segurada)
    
    return valor_mensal, N_x, N_x_n, anuidade_ajustada, percentual_mensal

def calcular_percentual_mensal(valor_mensal, taxa_juros, fracionamento, periodo, soma_segurada):
    taxa_fracionada = (1 + taxa_juros)**(1/fracionamento) - 1
    num_pagamentos = fracionamento * periodo
    
    if taxa_fracionada == 0:
        pgto = soma_segurada / num_pagamentos
    else:
        pgto = soma_segurada * taxa_fracionada * (1 + taxa_fracionada)**num_pagamentos / ((1 + taxa_fracionada)**num_pagamentos - 1)
    
    if pgto == 0:
        return 0, taxa_fracionada, num_pagamentos, pgto
    
    percentual = valor_mensal / pgto
    return percentual, taxa_fracionada, num_pagamentos, pgto

def calcular_saldo_devedor_price(soma_segurada, taxa_mensal, num_parcelas, periodo_t):
    """
    Calcula o saldo devedor de um financiamento Price no momento t.
    
    Args:
        soma_segurada: Valor inicial do financiamento (S₀)
        taxa_mensal: Taxa de juros mensal (i)
        num_parcelas: Número total de parcelas (n)
        periodo_t: Momento atual (t)
    
    Returns:
        Saldo devedor no momento t
    """
    if taxa_mensal == 0:
        return soma_segurada * (1 - periodo_t / num_parcelas)
    
    # Fórmula do saldo devedor Price
    saldo_devedor = soma_segurada * ((1 + taxa_mensal)**num_parcelas - (1 + taxa_mensal)**periodo_t) / ((1 + taxa_mensal)**num_parcelas - 1)
    return max(0, saldo_devedor)  # Não pode ser negativo

def calcular_reserva_matematica_prestamista(tabua_obj, idade, sexo, periodo, taxa_juros, soma_segurada, tempo_t, premio_mensal=None):
    """
    Calcula a reserva matemática (provisão prospectiva) do seguro prestamista no tempo t.
    
    A reserva matemática no tempo t representa o valor que a seguradora deve manter
    em reserva para honrar os compromissos futuros do seguro, considerando que o
    segurado sobreviveu até o tempo t.
    
    Args:
        tabua_obj: Objeto da tábua de comutação
        idade: Idade inicial do segurado
        sexo: Sexo do segurado ('M' ou 'F')
        periodo: Período total do seguro em meses
        taxa_juros: Taxa de juros anual
        soma_segurada: Soma segurada inicial
        tempo_t: Tempo t (0 <= t <= periodo) em meses
    
    Returns:
        Dicionário com a reserva matemática e detalhes do cálculo
    """
    # Validação do tempo t
    if tempo_t < 0 or tempo_t > periodo:
        raise ValueError(f"O tempo t deve estar entre 0 e {periodo} meses.")
    
    # Se t = 0, a reserva é o prêmio único
    if tempo_t == 0:
        resultado_prestamista = calcular_seguro_prestamista(tabua_obj, idade, sexo, periodo, taxa_juros, soma_segurada)
        return {
            "reserva_matematica": resultado_prestamista['premio_unico'],
            "explicacao": "No tempo t=0, a reserva matemática é igual ao prêmio único pago.",
            "detalhes": {
                "tempo_t": 0,
                "idade_atual": idade,
                "periodo_restante": periodo,
                "saldo_devedor_atual": soma_segurada,
                "valor_segurado_atual": soma_segurada
            }
        }
    
    # Se t = periodo, a reserva é zero (seguro terminou)
    if tempo_t == periodo:
        return {
            "reserva_matematica": 0.0,
            "explicacao": "No tempo t=período, o seguro terminou e a reserva matemática é zero.",
            "detalhes": {
                "tempo_t": tempo_t,
                "idade_atual": idade + tempo_t / 12,
                "periodo_restante": 0,
                "saldo_devedor_atual": 0.0,
                "valor_segurado_atual": 0.0
            }
        }
    
    # Configurar tábua para o sexo correto
    tabua_obj.dados = tabua_obj.tabuas_disponiveis[tabua_obj.tabua_selecionada]
    tabua_obj.calcular_tabua_comutacao()
    
    # Parâmetros do financiamento
    taxa_mensal = (1 + taxa_juros)**(1/12) - 1
    num_parcelas = periodo
    v_mensal = 1 / (1 + taxa_mensal)
    
    # Calcular PMT do financiamento
    if taxa_mensal == 0:
        pmt_financiamento = soma_segurada / num_parcelas
    else:
        pmt_financiamento = soma_segurada * taxa_mensal * (1 + taxa_mensal)**num_parcelas / ((1 + taxa_mensal)**num_parcelas - 1)
    
    # Calcular saldo devedor no tempo t
    saldo_devedor_t = calcular_saldo_devedor_price(soma_segurada, taxa_mensal, num_parcelas, tempo_t)
    
    # Idade atual no tempo t
    idade_atual = idade + tempo_t / 12
    periodo_restante = periodo - tempo_t
    
    # Cache de probabilidades mensais
    dados_sexo = tabua_obj.dados['masculino'] if sexo == 'M' else tabua_obj.dados['feminino']
    cache_qx_mensal = {}
    
    # Pré-calcular probabilidades de morte mensais para o período restante
    for k in range(1, int(periodo_restante) + 1):
        idade_k = int(idade_atual + k / 12)
        if idade_k in dados_sexo:
            qx_anual = tabua_obj.obter_qx(idade_k, sexo)
            cache_qx_mensal[idade_k] = 1 - (1 - qx_anual)**(1/12)
    
    # Calcular probabilidade de sobrevivência até o tempo t
    prob_sobrevivencia_t = 1.0
    for k in range(1, int(tempo_t) + 1):
        idade_k_1 = int(idade + (k - 1) / 12)
        if idade_k_1 in dados_sexo:
            qx_anual = tabua_obj.obter_qx(idade_k_1, sexo)
            qx_mensal = 1 - (1 - qx_anual)**(1/12)
            prob_sobrevivencia_t *= (1 - qx_mensal)
    
    # Calcular reserva matemática usando a fórmula mais simples e direta
    # V_t = Σ_{k=t+1}^{n} v^{k-t} ⋅ SD_k ⋅ _{k-t|1}q_{x+t} - Σ_{k=t+1}^{n} v^{k-t} ⋅ P ⋅ _{k-t}p_{x+t}
    
    # Calcular o prêmio mensal do seguro prestamista (se não foi fornecido)
    if premio_mensal is None:
        resultado_prestamista = calcular_seguro_prestamista(tabua_obj, idade, sexo, periodo, taxa_juros, soma_segurada)
        premio_mensal = resultado_prestamista['premio_mensal']
    
    # Função para calcular probabilidade de sobrevivência mensal
    def calcular_prob_sobrevivencia_mensal(idade_inicial, meses):
        """Calcula a probabilidade de sobrevivência para 'meses' períodos mensais"""
        prob = 1.0
        for i in range(meses):
            idade_atual = idade_inicial + i / 12
            idade_int = int(idade_atual)
            if idade_int in dados_sexo:
                qx_anual = tabua_obj.obter_qx(idade_int, sexo)
                qx_mensal = 1 - (1 - qx_anual)**(1/12)
                prob *= (1 - qx_mensal)
        return max(0, min(1, prob))
    
    # Primeira soma: Σ v^{k-t} ⋅ SD_k ⋅ _{k-t|1}q_{x+t} (benefícios)
    valor_presente_beneficios = 0.0
    
    # Segunda soma: Σ v^{k-t} ⋅ P ⋅ _{k-t}p_{x+t} (prêmios)
    valor_presente_premios = 0.0
    
    detalhes_calculo = []
    
    for k in range(int(tempo_t) + 1, int(periodo) + 1):
        # SD_k = Saldo devedor no mês k (benefício)
        SD_k = calcular_saldo_devedor_price(soma_segurada, taxa_mensal, num_parcelas, k)
        
        # Fator de desconto v^{k-t}
        v_power = v_mensal ** (k - tempo_t)
        
        # _{k-t}p_{x+t} - probabilidade de sobrevivência por k-t períodos
        k_t_p_x_t = calcular_prob_sobrevivencia_mensal(idade + tempo_t/12, k - tempo_t)
        
        # _{k-t|1}q_{x+t} - probabilidade de sobreviver k-t períodos e morrer no próximo
        # = _{k-t}p_{x+t} - _{k-t+1}p_{x+t}
        k_t_1_p_x_t = calcular_prob_sobrevivencia_mensal(idade + tempo_t/12, k - tempo_t + 1)
        k_t_1_q_x_t = max(0, k_t_p_x_t - k_t_1_p_x_t)
        
        # Contribuição para benefícios: v^{k-t} ⋅ SD_k ⋅ _{k-t|1}q_{x+t}
        contribuicao_beneficios = v_power * SD_k * k_t_1_q_x_t
        valor_presente_beneficios += contribuicao_beneficios
        
        # Contribuição para prêmios: v^{k-t} ⋅ P ⋅ _{k-t}p_{x+t}
        contribuicao_premios = v_power * premio_mensal * k_t_p_x_t
        valor_presente_premios += contribuicao_premios
        
        # Detalhes do cálculo
        detalhes_calculo.append({
            "mes": k,
            "idade": idade + k / 12,
            "saldo_devedor": SD_k,
            "premio": premio_mensal,
            "k_t_p_x_t": k_t_p_x_t,
            "k_t_1_q_x_t": k_t_1_q_x_t,
            "fator_desconto": v_power,
            "contribuicao_beneficios": contribuicao_beneficios,
            "contribuicao_premios": contribuicao_premios
        })
    
    # V_t = VP(Benefícios) - VP(Prêmios)
    reserva_matematica = valor_presente_beneficios - valor_presente_premios
    
    # Debug: imprimir valores para análise
    # Debug removido para limpeza do terminal
    # Debug removido para limpeza do terminal
    
    return {
        "reserva_matematica": reserva_matematica,
        "valor_presente_beneficios": valor_presente_beneficios,
        "valor_presente_premios": valor_presente_premios,
        "explicacao": f"Reserva matemática no tempo t={tempo_t} meses, considerando que o segurado sobreviveu até este momento.",
        "detalhes": {
            "tempo_t": tempo_t,
            "idade_atual": idade_atual,
            "periodo_restante": periodo_restante,
            "saldo_devedor_atual": saldo_devedor_t,
            "valor_segurado_atual": saldo_devedor_t,
            "prob_sobrevivencia_t": prob_sobrevivencia_t,
            "calculo_detalhado": detalhes_calculo
        }
    }

def calcular_seguro_prestamista(tabua_obj, idade, sexo, periodo, taxa_juros, soma_segurada):
    """
    Calcula o seguro prestamista onde o valor segurado diminui seguindo a amortização Price.
    VERSÃO SUPER OTIMIZADA - Máxima performance com O(n) linear
    
    Args:
        tabua_obj: Objeto da tábua de comutação
        idade: Idade inicial do segurado
        sexo: Sexo do segurado ('M' ou 'F')
        periodo: Período do seguro em meses
        taxa_juros: Taxa de juros anual
        soma_segurada: Soma segurada inicial
    
    Returns:
        Dicionário com os resultados do cálculo
    """
    # Configurar tábua para o sexo correto
    tabua_obj.dados = tabua_obj.tabuas_disponiveis[tabua_obj.tabua_selecionada]
    tabua_obj.calcular_tabua_comutacao()
    
    # Parâmetros do financiamento
    taxa_mensal = (1 + taxa_juros)**(1/12) - 1
    num_parcelas = periodo
    v_mensal = 1 / (1 + taxa_mensal)
    
    # Calcular PMT do financiamento
    if taxa_mensal == 0:
        pmt_financiamento = soma_segurada / num_parcelas
    else:
        pmt_financiamento = soma_segurada * taxa_mensal * (1 + taxa_mensal)**num_parcelas / ((1 + taxa_mensal)**num_parcelas - 1)
    
    # SUPER OTIMIZAÇÃO 1: Pré-calcular TUDO de uma vez
    # Saldos devedor, probabilidades, fatores de desconto
    saldos_devedor = [calcular_saldo_devedor_price(soma_segurada, taxa_mensal, num_parcelas, k) for k in range(num_parcelas + 1)]
    
    # SUPER OTIMIZAÇÃO 2: Cache otimizado de probabilidades mensais
    dados_sexo = tabua_obj.dados['masculino'] if sexo == 'M' else tabua_obj.dados['feminino']
    cache_qx_mensal = {}
    
    # Pré-calcular todas as idades necessárias
    idades_necessarias = set()
    for k in range(num_parcelas + 1):
        idades_necessarias.add(int(idade + k / 12))
    
    for idade_int in idades_necessarias:
        if idade_int in dados_sexo:
            qx_anual = tabua_obj.obter_qx(idade_int, sexo)
            cache_qx_mensal[idade_int] = 1 - (1 - qx_anual)**(1/12)
    
    # SUPER OTIMIZAÇÃO 3: Cálculo vetorizado de probabilidades de sobrevivência
    prob_sobrevivencias = [1.0]  # Inicia com 100%
    prob_atual = 1.0
    
    # SUPER OTIMIZAÇÃO 4: Cache de fatores de desconto
    v_powers = [v_mensal ** k for k in range(1, num_parcelas + 1)]
    
    # SUPER OTIMIZAÇÃO 5: Cache de lx para anuidade
    lx_0 = tabua_obj.l_x.get(idade, 1)
    lx_cache = {}
    for t in range(1, num_parcelas + 1):
        idade_t = int(idade + t / 12)
        if idade_t in tabua_obj.l_x:
            lx_cache[t] = tabua_obj.l_x[idade_t]
    
    # Calcular prêmio único do seguro prestamista
    premio_unico = 0
    detalhes_calculo = []
    
    # SUPER OTIMIZAÇÃO 6: Loop principal otimizado
    for k in range(1, num_parcelas + 1):
        # Usar arrays pré-calculados
        saldo_devedor = saldos_devedor[k - 1]
        fator_desconto = v_powers[k - 1]
        
        # SUPER OTIMIZAÇÃO 7: Cálculo otimizado de juros e amortização
        if k == 1:
            juros_mes = soma_segurada * taxa_mensal
        else:
            juros_mes = saldos_devedor[k - 2] * taxa_mensal
        
        amortizacao_mes = pmt_financiamento - juros_mes
        
        # SUPER OTIMIZAÇÃO 8: Cálculo incremental de sobrevivência
        if k > 1:
            idade_anterior = int(idade + (k - 2) / 12)
            if idade_anterior in cache_qx_mensal:
                prob_atual *= (1 - cache_qx_mensal[idade_anterior])
        
        prob_sobrevivencias.append(prob_atual)
        
        # SUPER OTIMIZAÇÃO 9: Lookup otimizado de probabilidade de morte
        idade_k_1 = int(idade + (k - 1) / 12)
        qx_k = cache_qx_mensal.get(idade_k_1, 1.0)
        
        # Contribuição para o prêmio único (otimizada)
        contribuicao = saldo_devedor * prob_atual * qx_k * fator_desconto
        premio_unico += contribuicao
        
        # Detalhes (apenas se necessário)
        detalhes_calculo.append({
            'periodo': k,
            'idade': idade_k_1,
            'saldo_devedor': saldo_devedor,
            'pmt': pmt_financiamento,
            'juros': juros_mes,
            'amortizacao': amortizacao_mes,
            'prob_morte': qx_k,
            'fator_desconto': fator_desconto,
            'valor_antes_desconto': saldo_devedor * prob_atual * qx_k,
            'contribuicao': contribuicao
        })
    
    # SUPER OTIMIZAÇÃO 10: Cálculo vetorizado da anuidade mensal
    anuidade_mensal = 0
    v_power = v_mensal
    
    for t in range(1, num_parcelas + 1):
        if t in lx_cache and lx_0 > 0:
            prob_sobrevivencia = lx_cache[t] / lx_0
            anuidade_mensal += prob_sobrevivencia * v_power
        v_power *= v_mensal
    
    if anuidade_mensal > 0:
        premio_mensal = premio_unico / anuidade_mensal
    else:
        premio_mensal = 0
    
    # Calcular percentual mensal
    percentual_mensal = (premio_mensal / pmt_financiamento) if pmt_financiamento > 0 else 0
    
    # Calcular soma total das contribuições
    soma_contribuicoes = sum(item['contribuicao'] for item in detalhes_calculo)
    
    # SUPER OTIMIZAÇÃO 11: Taxa de quitação otimizada
    pv_liquido = -(soma_segurada - premio_unico)
    
    # Método de Newton-Raphson otimizado
    taxa_aprox = 0.01
    for i in range(3):  # Reduzido de 5 para 3 iterações
        npv = pv_liquido
        for t in range(1, num_parcelas + 1):
            npv += pmt_financiamento / ((1 + taxa_aprox) ** t)
        
        if abs(npv) < 0.01:
            break
            
        # Derivada numérica otimizada
        npv_plus = pv_liquido
        for t in range(1, num_parcelas + 1):
            npv_plus += pmt_financiamento / ((1 + taxa_aprox + 0.001) ** t)
        
        derivada = (npv_plus - npv) / 0.001
        if abs(derivada) > 1e-10:
            taxa_aprox = taxa_aprox - npv / derivada
        else:
            break
    
    taxa_quitação_risco_mensal = (1 + taxa_aprox) / (1 + taxa_mensal) - 1
    
    return {
        'premio_unico': premio_unico,
        'premio_mensal': premio_mensal,
        'percentual_mensal': percentual_mensal,
        'pmt_financiamento': pmt_financiamento,
        'detalhes_calculo': detalhes_calculo,
        'taxa_mensal': taxa_mensal,
        'num_parcelas': num_parcelas,
        'soma_contribuicoes': soma_contribuicoes,
        'taxa_quitação_risco_mensal': taxa_quitação_risco_mensal
    }

def calcular_seguro_prestamista_alt(tabua_obj, idade, sexo, periodo, taxa_juros, soma_segurada):
    """
    Calcula o seguro prestamista usando metodologia alternativa.
    
    Metodologia:
    - lx: probabilidade de sobrevivência mensal (interpolação linear da tábua anual)
    - qx: probabilidade de morte mensal
    - vx: fator de desconto mensal
    - PGTO: parcela mensal do financiamento (Price)
    - VP PGTO: PGTO * vx
    - VPA PGTO: VP PGTO * lx
    - Prêmio único = Soma(VP PGTO) - Soma(VPA PGTO)
    - Taxa cobertura risco = TAXA(nper, pgto, -vp) ajustada
    
    Args:
        tabua_obj: Objeto da tábua de comutação
        idade: Idade inicial do segurado
        sexo: Sexo do segurado ('M' ou 'F')
        periodo: Período do seguro em meses
        taxa_juros: Taxa de juros anual
        soma_segurada: Soma segurada inicial
    
    Returns:
        Dicionário com os resultados do cálculo
    """
    # Configurar tábua para o sexo correto
    tabua_obj.dados = tabua_obj.tabuas_disponiveis[tabua_obj.tabua_selecionada]
    tabua_obj.calcular_tabua_comutacao()
    
    # Parâmetros do financiamento
    taxa_mensal = (1 + taxa_juros)**(1/12) - 1
    num_parcelas = periodo
    v_mensal = 1 / (1 + taxa_mensal)  # Fator de desconto mensal
    
    # Calcular PMT do financiamento (Price)
    if taxa_mensal == 0:
        pmt_financiamento = soma_segurada / num_parcelas
    else:
        pmt_financiamento = soma_segurada * taxa_mensal * (1 + taxa_mensal)**num_parcelas / ((1 + taxa_mensal)**num_parcelas - 1)
    
    # Inicializar variáveis para cálculos
    soma_vp_pgto = 0  # Soma da coluna VP PGTO
    soma_vpa_pgto = 0  # Soma da coluna VPA PGTO
    detalhes_calculo = []
    
    # Calcular lx inicial (probabilidade de sobrevivência na idade inicial)
    lx_inicial = 1.0
    
    for k in range(num_parcelas + 1):  # k vai de 0 até n (incluindo tempo 0)
        # Idade no momento k (idade fracionária)
        idade_k = idade + k / 12
        
        # Calcular lx (probabilidade de sobrevivência até o momento k)
        if k == 0:
            lx_k = 1.0
        else:
            # Interpolação linear da tábua anual para mensal
            lx_k = calcular_lx_mensal(tabua_obj, idade, sexo, k)
        
        # Calcular qx (probabilidade de morte no mês k+1, condicional em estar vivo no início do mês k+1)
        if k < num_parcelas:
            # Interpolação linear da tábua anual para mensal
            qx_k = calcular_qx_mensal(tabua_obj, idade, sexo, k)
        else:
            qx_k = 0  # Não há morte após o último período
        
        # Calcular vx (fator de desconto para o momento k)
        vx_k = v_mensal ** k
        
        # Calcular PGTO (parcela do financiamento)
        if k == 0:
            pgto_k = 0  # Não há pagamento no tempo 0
        else:
            pgto_k = pmt_financiamento
        
        # Calcular VP PGTO (Valor Presente da Parcela)
        vp_pgto_k = pgto_k * vx_k
        
        # Calcular VPA PGTO (Valor Presente Atuarial da Parcela)
        vpa_pgto_k = vp_pgto_k * lx_k
        
        # Acumular somas
        soma_vp_pgto += vp_pgto_k
        soma_vpa_pgto += vpa_pgto_k
        
        # Saldo devedor no momento k (para referência)
        if k == 0:
            saldo_devedor_k = soma_segurada
        else:
            saldo_devedor_k = calcular_saldo_devedor_price(soma_segurada, taxa_mensal, num_parcelas, k)
        
        # Detalhes do cálculo
        detalhes_calculo.append({
            'tempo': k,
            'idade': int(idade_k),
            'lx': lx_k,
            'qx': qx_k,
            'vx': vx_k,
            'pgto': pgto_k,
            'vp_pgto': vp_pgto_k,
            'vpa_pgto': vpa_pgto_k,
            'saldo_devedor': saldo_devedor_k
        })
    
    # Calcular prêmio único (diferença entre VP PGTO e VPA PGTO)
    premio_unico = soma_vp_pgto - soma_vpa_pgto
    
    # Calcular Taxa de Cobertura de Risco primeiro
    # Fórmula: (1 + TAXA(nper=PrazoFinanciamento, pgto=Parcela, -vp=SomaColunaVPAPGTO)) / (1+JurosMensal) - 1
    try:
        import numpy as np
        from scipy.optimize import fsolve
        
        # Valor presente líquido = -soma_vpa_pgto (negativo para TAXA)
        pv_liquido = -soma_vpa_pgto
        
        # Função para calcular NPV
        def npv_func(rate):
            npv = pv_liquido
            for t in range(1, num_parcelas + 1):
                npv += pmt_financiamento / ((1 + rate) ** t)
            return npv
        
        # Encontrar a taxa que zera o NPV
        taxa_implícita = fsolve(npv_func, 0.01)[0]
        
        # Ajustar pela taxa de juros mensal
        taxa_cobertura_risco = (1 + taxa_implícita) / (1 + taxa_mensal) - 1
        
    except Exception as e:
        # Fallback se scipy não estiver disponível
        taxa_cobertura_risco = 0.0
    
    # Calcular prêmio mensal usando a fórmula especificada
    # PGTO((1+JurosMensal)*(1+TaxaCoberturaRisco)-1; PrazoMeses; -PremioUnico)
    taxa_efetiva = (1 + taxa_mensal) * (1 + taxa_cobertura_risco) - 1
    
    if taxa_efetiva == 0:
        premio_mensal = premio_unico / num_parcelas
    else:
        premio_mensal = premio_unico * taxa_efetiva * (1 + taxa_efetiva)**num_parcelas / ((1 + taxa_efetiva)**num_parcelas - 1)
    
    # Calcular percentual mensal em relação ao PMT do financiamento
    percentual_mensal = (premio_mensal / pmt_financiamento) if pmt_financiamento > 0 else 0
    
    
    return {
        'premio_unico': premio_unico,
        'premio_mensal': premio_mensal,
        'percentual_mensal': percentual_mensal,
        'pmt_financiamento': pmt_financiamento,
        'detalhes_calculo': detalhes_calculo,
        'taxa_mensal': taxa_mensal,
        'num_parcelas': num_parcelas,
        'soma_vp_pgto': soma_vp_pgto,
        'soma_vpa_pgto': soma_vpa_pgto,
        'taxa_cobertura_risco': taxa_cobertura_risco,
        'valor_emprestimo_atual': soma_vpa_pgto  # Valor do empréstimo atuarial
    }

def calcular_lx_mensal(tabua_obj, idade_inicial, sexo, k):
    """
    Calcula a probabilidade de sobrevivência mensal de forma cumulativa.
    
    l0 = 1
    l1 = l0 * (1 - q0)
    l2 = l1 * (1 - q1)
    l3 = l2 * (1 - q2)
    ...
    
    Args:
        tabua_obj: Objeto da tábua de comutação
        idade_inicial: Idade inicial
        sexo: Sexo ('M' ou 'F')
        k: Número de meses (0, 1, 2, ...)
    
    Returns:
        Probabilidade de sobrevivência até o mês k
    """
    if k == 0:
        return 1.0
    
    # Obter dados da tábua
    dados_sexo = tabua_obj.dados['masculino'] if sexo == 'M' else tabua_obj.dados['feminino']
    
    # Calcular lx de forma cumulativa
    lx = 1.0  # l0 = 1
    
    for i in range(k):
        # Calcular idade no mês i
        # Do tempo 0 ao 11 = idade atual, do tempo 12 ao 23 = próxima idade, etc.
        idade_i = idade_inicial + i // 12
        idade_inteira_i = int(idade_i)
        
        # Obter qx anual para a idade inteira
        if idade_inteira_i in dados_sexo:
            qx_anual = dados_sexo[idade_inteira_i]
            # Converter de anual para mensal
            qx_mensal = 1 - (1 - qx_anual)**(1/12)
            # Aplicar probabilidade de sobrevivência
            lx *= (1 - qx_mensal)
        else:
            lx = 0.0
            break
    
    return lx

def calcular_qx_mensal(tabua_obj, idade_inicial, sexo, k):
    """
    Calcula a probabilidade de morte mensal.
    O qx é constante para a mesma idade (mesmo ano).
    
    Args:
        tabua_obj: Objeto da tábua de comutação
        idade_inicial: Idade inicial
        sexo: Sexo ('M' ou 'F')
        k: Número de meses (0, 1, 2, ...)
    
    Returns:
        Probabilidade de morte no mês k+1
    """
    # Calcular a idade no início do mês k+1
    idade_k_1 = idade_inicial + k / 12
    idade_inteira = int(idade_k_1)
    
    # Obter qx anual para a idade inteira
    dados_sexo = tabua_obj.dados['masculino'] if sexo == 'M' else tabua_obj.dados['feminino']
    
    if idade_inteira in dados_sexo:
        qx_anual = dados_sexo[idade_inteira]
        # Converter de anual para mensal usando aproximação atuarial
        qx_mensal = 1 - (1 - qx_anual)**(1/12)
        return qx_mensal
    else:
        return 1.0  # Morte certa após idade máxima

@lru_cache(maxsize=1000)
def calcular_taxas_seguro_cached(tabua_nome, idade, sexo, periodo, taxa_juros, soma_segurada=100000):
    """
    Versão otimizada com cache da função calcular_taxas_seguro.
    Cache baseado nos parâmetros de entrada para evitar recálculos.
    """
    try:
        # Usar cache global para tábuas
        cache_key = f"{tabua_nome}_{taxa_juros}_{sexo}"
        
        if cache_key not in TABUAS_CACHE:
            # Criar nova instância da tábua
            tabua_obj = TabuladeComutacao(taxa_juros, tabua_nome)
            tabua_obj.tabua_selecionada = tabua_nome
            tabua_obj.dados = tabua_obj.tabuas_disponiveis[tabua_nome]
            tabua_obj.calcular_tabua_comutacao()
            TABUAS_CACHE[cache_key] = tabua_obj
        else:
            tabua_obj = TABUAS_CACHE[cache_key]
        
        # Calcular seguro anual
        seguro_anual = calcular_seguro_anual(tabua_obj, idade, sexo, periodo)
        
        # Calcular seguro fracionado total (taxa à vista)
        seguro_fracionado_total, _, _ = calcular_seguro_fracionado_total(taxa_juros, 12, seguro_anual)
        
        # Calcular valor monetário à vista
        valor_monetario_vista = soma_segurada * seguro_fracionado_total
        
        # Calcular prêmio mensal
        valor_mensal, _, _, _, percentual_mensal_calc = calcular_premio_mensal(
            tabua_obj, idade, periodo, valor_monetario_vista, taxa_juros, 12, soma_segurada
        )
        
        return {
            'taxa_vista': seguro_fracionado_total,
            'taxa_mensal': percentual_mensal_calc[0] if percentual_mensal_calc else 0
        }
        
    except Exception as e:
        return {
            'taxa_vista': 0.0,
            'taxa_mensal': 0.0
        }

def calcular_taxas_seguro(tabua_obj, idade, sexo, periodo, taxa_juros, soma_segurada=100000):
    """
    Wrapper para manter compatibilidade com código existente.
    Agora usa a versão otimizada com cache.
    """
    return calcular_taxas_seguro_cached(
        tabua_obj.tabua_selecionada, idade, sexo, periodo, taxa_juros, soma_segurada
    )

def processar_combinacao_paralela(args):
    """
    Função para processar uma única combinação em paralelo.
    Args: (tabua_nome, idade, sexo, periodo, taxa_juros, tipo_tabua, soma_segurada)
    """
    try:
        tabua_nome, idade, sexo, periodo, taxa_juros, tipo_tabua, soma_segurada = args
        
        # Calcular taxas usando a versão otimizada
        taxas = calcular_taxas_seguro_cached(tabua_nome, idade, sexo, periodo, taxa_juros, soma_segurada)
        
        return {
            "idade": idade,
            "sexo": sexo,
            "periodo": periodo,
            "tipo_tabua": tipo_tabua,
            "tabua": tabua_nome,
            "taxa_vista": f"{taxas['taxa_vista']*100:.4f}%",
            "taxa_mensal": f"{taxas['taxa_mensal']*100:.4f}%"
        }
    except Exception as e:
        return None

def calcular_coletivo_paralelo(idade_min, idade_max, sexos, periodo_min, periodo_max, 
                              taxa_juros, tabuas_validas, tabuas_invalidas, max_workers=None):
    """
    Versão otimizada do cálculo coletivo usando paralelismo.
    """
    if max_workers is None:
        max_workers = min(mp.cpu_count(), 8)  # Limitar a 8 workers para não sobrecarregar
    
    # Preparar todas as combinações
    combinacoes = []
    
    # Adicionar combinações para tábuas válidas
    for tabua in tabuas_validas:
        for idade in range(idade_min, idade_max + 1):
            for sexo in sexos:
                for periodo in range(periodo_min, periodo_max + 1):
                    combinacoes.append((tabua, idade, sexo, periodo, taxa_juros, "Válido", 100000))
    
    # Adicionar combinações para tábuas inválidas
    for tabua in tabuas_invalidas:
        for idade in range(idade_min, idade_max + 1):
            for sexo in sexos:
                for periodo in range(periodo_min, periodo_max + 1):
                    combinacoes.append((tabua, idade, sexo, periodo, taxa_juros, "Inválido", 100000))
    
    print(f"Processando {len(combinacoes)} combinacoes com {max_workers} workers paralelos...")
    inicio = time.time()
    
    # Processar em paralelo
    resultados = []
    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        # Submeter todas as tarefas
        future_to_combinacao = {
            executor.submit(processar_combinacao_paralela, comb): comb 
            for comb in combinacoes
        }
        
        # Coletar resultados conforme completam
        for i, future in enumerate(as_completed(future_to_combinacao)):
            resultado = future.result()
            if resultado is not None:
                resultados.append(resultado)
            
            # Mostrar progresso
            if (i + 1) % 50 == 0 or i == len(combinacoes) - 1:
                progresso = ((i + 1) / len(combinacoes)) * 100
                print(f"Progresso: {progresso:.1f}% ({i + 1}/{len(combinacoes)})")
    
    fim = time.time()
    tempo_total = fim - inicio
    print(f"Processamento concluido em {tempo_total:.2f} segundos")
    print(f"Velocidade: {len(combinacoes)/tempo_total:.1f} combinacoes/segundo")
    
    return resultados

def limpar_cache_tabuas():
    """Limpa o cache de tábuas para liberar memória."""
    global TABUAS_CACHE
    TABUAS_CACHE.clear()
    calcular_taxas_seguro_cached.cache_clear()
    print("🧹 Cache de tábuas limpo")

def obter_estatisticas_cache():
    """Retorna estatísticas do cache."""
    return {
        "tabulas_em_cache": len(TABUAS_CACHE),
        "cache_hits": calcular_taxas_seguro_cached.cache_info().hits,
        "cache_misses": calcular_taxas_seguro_cached.cache_info().misses,
        "tamanho_cache": calcular_taxas_seguro_cached.cache_info().currsize
    }

class CalculadoraHandler(http.server.SimpleHTTPRequestHandler):
    def do_GET(self):
        if self.path == '/':
            self.path = '/index.html'
            return super().do_GET()
        elif self.path == '/index.html':
            return super().do_GET()
        elif self.path == '/calculadora_individual.html':
            return super().do_GET()
        elif self.path == '/calculadora_coletiva.html':
            return super().do_GET()
        elif self.path == '/tabuas':
            self.obter_tabuas_disponiveis()
        elif self.path == '/download_excel':
            self.handle_download_excel()
        elif self.path == '/cache_stats':
            self.handle_cache_stats()
        elif self.path == '/limpar_cache':
            self.handle_limpar_cache()
        elif self.path == '/calcular_prestamista':
            self.handle_calcular_prestamista()
        elif self.path == '/calcular_prestamista_alt':
            self.handle_calcular_prestamista_alt()
        else:
            return super().do_GET()
    
    def obter_tabuas_disponiveis(self):
        try:
            # Criar instância temporária para obter tábuas
            tabua_obj = TabuladeComutacao(0.1)  # Taxa temporária
            tabuas = list(tabua_obj.tabuas_disponiveis.keys())
            
            response = {
                "success": True,
                "tabuas": tabuas,
                "tabua_padrao": tabua_obj.tabua_padrao
            }
            
            self.send_response(200)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(response, ensure_ascii=False).encode('utf-8'))
            
        except Exception as e:
            error_response = {"success": False, "error": str(e)}
            self.send_response(500)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(error_response, ensure_ascii=False).encode('utf-8'))
    
    def end_headers(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        super().end_headers()
    
    def do_OPTIONS(self):
        self.send_response(200)
        self.end_headers()
    
    def handle_calcular_coletivo(self):
        try:
            # Ler dados do POST
            content_length = int(self.headers.get('Content-Length', 0))
            if content_length == 0:
                raise ValueError("Content-Length is 0")
            
            post_data = self.rfile.read(content_length)
            if not post_data:
                raise ValueError("No data received")
            
            data = json.loads(post_data.decode('utf-8'))
            
            # Extrair parâmetros
            idade_min = int(data['idade_min'])
            idade_max = int(data['idade_max'])
            sexos = data['sexos']
            periodo_min = int(data['periodo_min'])
            periodo_max = int(data['periodo_max'])
            taxa_juros = float(data['taxa_juros']) / 100
            tabuas_validas = data['tabuas_validas']
            tabuas_invalidas = data['tabuas_invalidas']
            
            # Validação dos limites
            if not (0 <= idade_min <= idade_max <= 110):
                raise ValueError("As idades devem estar entre 0 e 110 anos, e a idade mínima deve ser menor ou igual à máxima.")
            
            if not (1 <= periodo_min <= periodo_max <= 10):
                raise ValueError("Os períodos devem estar entre 1 e 120 meses, e o período mínimo deve ser menor ou igual ao máximo.")
            
            if not (0 <= taxa_juros <= 0.20):  # 0% a 20%
                raise ValueError("A taxa de juros deve estar entre 0% e 20%.")
            
            if len(sexos) == 0:
                raise ValueError("Selecione pelo menos um sexo.")
            
            if len(tabuas_validas) == 0 and len(tabuas_invalidas) == 0:
                raise ValueError("Selecione pelo menos uma tábua válida ou inválida.")
            
            # Calcular total de combinações
            total_idades = idade_max - idade_min + 1
            total_periodos = periodo_max - periodo_min + 1
            total_sexos = len(sexos)
            total_tabuas = len(tabuas_validas) + len(tabuas_invalidas)
            total_combinacoes = total_idades * total_periodos * total_sexos * total_tabuas
            
            print(f"Iniciando calculo coletivo otimizado:")
            print(f"   • Idades: {idade_min}-{idade_max} ({total_idades} idades)")
            print(f"   • Períodos: {periodo_min}-{periodo_max} ({total_periodos} períodos)")
            print(f"   • Sexos: {sexos} ({total_sexos} sexos)")
            print(f"   • Tábuas: {len(tabuas_validas)} válidas + {len(tabuas_invalidas)} inválidas")
            print(f"   • Total: {total_combinacoes} combinações")
            
            # Usar versão paralela otimizada
            resultados = calcular_coletivo_paralelo(
                idade_min, idade_max, sexos, periodo_min, periodo_max,
                taxa_juros, tabuas_validas, tabuas_invalidas
            )
            
            # Filtrar resultados válidos
            resultados_validos = [r for r in resultados if r is not None]
            
            # Preparar resposta
            response = {
                "success": True,
                "resultados": resultados_validos,
                "total_combinacoes": total_combinacoes,
                "combinacoes_processadas": len(resultados_validos),
                "progresso": 100.0,
                "idade_min": idade_min,
                "idade_max": idade_max,
                "periodo_min": periodo_min,
                "periodo_max": periodo_max,
                "tabuas_utilizadas": list(set(tabuas_validas + tabuas_invalidas)),
                "otimizado": True,
                "cache_hits": len(TABUAS_CACHE)
            }
            
            # Enviar resposta
            self.send_response(200)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(response, ensure_ascii=False).encode('utf-8'))
            
        except Exception as e:
            error_response = {"success": False, "error": str(e)}
            self.send_response(500)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(error_response, ensure_ascii=False).encode('utf-8'))
    
    def handle_calcular_coletivo_progress(self):
        try:
            # Ler dados do POST
            content_length = int(self.headers.get('Content-Length', 0))
            if content_length == 0:
                raise ValueError("Content-Length is 0")
            
            post_data = self.rfile.read(content_length)
            if not post_data:
                raise ValueError("No data received")
            
            data = json.loads(post_data.decode('utf-8'))
            
            # Extrair parâmetros
            idade_min = int(data['idade_min'])
            idade_max = int(data['idade_max'])
            sexos = data['sexos']
            periodo_min = int(data['periodo_min'])
            periodo_max = int(data['periodo_max'])
            taxa_juros = float(data['taxa_juros']) / 100
            tabuas_validas = data['tabuas_validas']
            tabuas_invalidas = data['tabuas_invalidas']
            
            # Validação dos limites
            if not (0 <= idade_min <= idade_max <= 110):
                raise ValueError("As idades devem estar entre 0 e 110 anos, e a idade mínima deve ser menor ou igual à máxima.")
            
            if not (1 <= periodo_min <= periodo_max <= 10):
                raise ValueError("Os períodos devem estar entre 1 e 120 meses, e o período mínimo deve ser menor ou igual ao máximo.")
            
            if not (0 <= taxa_juros <= 0.20):  # 0% a 20%
                raise ValueError("A taxa de juros deve estar entre 0% e 20%.")
            
            if len(sexos) == 0:
                raise ValueError("Selecione pelo menos um sexo.")
            
            if len(tabuas_validas) == 0 and len(tabuas_invalidas) == 0:
                raise ValueError("Selecione pelo menos uma tábua válida ou inválida.")
            
            # Calcular total de combinações para progresso
            total_idades = idade_max - idade_min + 1
            total_periodos = periodo_max - periodo_min + 1
            total_sexos = len(sexos)
            total_tabuas = len(tabuas_validas) + len(tabuas_invalidas)
            total_combinacoes = total_idades * total_periodos * total_sexos * total_tabuas
            
            # Configurar Server-Sent Events
            self.send_response(200)
            self.send_header('Content-Type', 'text/event-stream')
            self.send_header('Cache-Control', 'no-cache')
            self.send_header('Connection', 'keep-alive')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            
            # Calcular todas as combinações
            resultados = []
            tabuas_utilizadas = set()
            combinacoes_processadas = 0
            
            # Processar tábuas válidas
            for tabua in tabuas_validas:
                tabuas_utilizadas.add(tabua)
                for idade in range(idade_min, idade_max + 1):
                    for sexo in sexos:
                        for periodo in range(periodo_min, periodo_max + 1):
                            try:
                                # Calcular seguro usando a nova função unificada
                                tabua_obj = TabuladeComutacao(taxa_juros, tabua)
                                tabua_obj.tabua_selecionada = tabua
                                
                                if tabua not in tabua_obj.tabuas_disponiveis:
                                    # Se a tábua não for encontrada, usar AT-83 como padrão
                                    if 'AT-83' in tabua_obj.tabuas_disponiveis:
                                        tabua = 'AT-83'
                                        tabua_obj.tabua_selecionada = tabua
                                    elif tabua_obj.tabuas_disponiveis:
                                        tabua = list(tabua_obj.tabuas_disponiveis.keys())[0]
                                        tabua_obj.tabua_selecionada = tabua
                                    else:
                                        continue
                                
                                # Usar a função unificada para calcular as taxas
                                taxas = calcular_taxas_seguro(tabua_obj, idade, sexo, periodo, taxa_juros, 100000)
                                
                                resultado = {
                                    "idade": idade,
                                    "sexo": sexo,
                                    "periodo": periodo,
                                    "tipo_tabua": "Válido",
                                    "tabua": tabua,
                                    "taxa_vista": f"{taxas['taxa_vista']*100:.4f}%",
                                    "taxa_mensal": f"{taxas['taxa_mensal']*100:.4f}%"
                                }
                                resultados.append(resultado)
                                
                            except Exception as e:
                                pass
                            
                            # Atualizar progresso e enviar via SSE
                            combinacoes_processadas += 1
                            progresso = (combinacoes_processadas / total_combinacoes) * 100
                            
                            # Enviar atualização de progresso
                            progress_data = {
                                "progresso": progresso,
                                "combinacoes_processadas": combinacoes_processadas,
                                "total_combinacoes": total_combinacoes,
                                "completo": False
                            }
                            
                            self.wfile.write(f"data: {json.dumps(progress_data)}\n\n".encode('utf-8'))
                            self.wfile.flush()
            
            # Processar tábuas inválidas
            for tabua in tabuas_invalidas:
                tabuas_utilizadas.add(tabua)
                for idade in range(idade_min, idade_max + 1):
                    for sexo in sexos:
                        for periodo in range(periodo_min, periodo_max + 1):
                            try:
                                # Calcular seguro usando a nova função unificada
                                tabua_obj = TabuladeComutacao(taxa_juros, tabua)
                                tabua_obj.tabua_selecionada = tabua
                                
                                if tabua not in tabua_obj.tabuas_disponiveis:
                                    # Se a tábua não for encontrada, usar AT-83 como padrão
                                    if 'AT-83' in tabua_obj.tabuas_disponiveis:
                                        tabua = 'AT-83'
                                        tabua_obj.tabua_selecionada = tabua
                                    elif tabua_obj.tabuas_disponiveis:
                                        tabua = list(tabua_obj.tabuas_disponiveis.keys())[0]
                                        tabua_obj.tabua_selecionada = tabua
                                    else:
                                        continue
                                
                                # Usar a função unificada para calcular as taxas
                                taxas = calcular_taxas_seguro(tabua_obj, idade, sexo, periodo, taxa_juros, 100000)
                                
                                resultado = {
                                    "idade": idade,
                                    "sexo": sexo,
                                    "periodo": periodo,
                                    "tipo_tabua": "Inválido",
                                    "tabua": tabua,
                                    "taxa_vista": f"{taxas['taxa_vista']*100:.4f}%",
                                    "taxa_mensal": f"{taxas['taxa_mensal']*100:.4f}%"
                                }
                                resultados.append(resultado)
                                
                            except Exception as e:
                                pass
                            
                            # Atualizar progresso e enviar via SSE
                            combinacoes_processadas += 1
                            progresso = (combinacoes_processadas / total_combinacoes) * 100
                            
                            # Enviar atualização de progresso
                            progress_data = {
                                "progresso": progresso,
                                "combinacoes_processadas": combinacoes_processadas,
                                "total_combinacoes": total_combinacoes,
                                "completo": False
                            }
                            
                            self.wfile.write(f"data: {json.dumps(progress_data)}\n\n".encode('utf-8'))
                            self.wfile.flush()
            
            # Enviar resultado final
            final_data = {
                "success": True,
                "resultados": resultados,
                "total_combinacoes": total_combinacoes,
                "combinacoes_processadas": combinacoes_processadas,
                "progresso": 100.0,
                "idade_min": idade_min,
                "idade_max": idade_max,
                "periodo_min": periodo_min,
                "periodo_max": periodo_max,
                "tabuas_utilizadas": list(tabuas_utilizadas),
                "completo": True
            }
            
            self.wfile.write(f"data: {json.dumps(final_data)}\n\n".encode('utf-8'))
            self.wfile.flush()
            
        except Exception as e:
            error_data = {
                "success": False,
                "error": f"Erro no cálculo coletivo: {str(e)}",
                "completo": True
            }
            self.wfile.write(f"data: {json.dumps(error_data)}\n\n".encode('utf-8'))
            self.wfile.flush()
    
    def handle_download_excel(self):
        try:
            # Ler dados do POST
            content_length = int(self.headers.get('Content-Length', 0))
            if content_length == 0:
                raise ValueError("Content-Length is 0")
            
            post_data = self.rfile.read(content_length)
            if not post_data:
                raise ValueError("No data received")
            
            data = json.loads(post_data.decode('utf-8'))
            resultados = data.get('resultados', [])
            
            if not resultados:
                raise ValueError("Nenhum resultado para exportar")
            
            # Criar arquivo Excel
            excel_data = self.create_excel_data(resultados)
            
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', 'attachment; filename="resultados_analise_coletiva.xlsx"')
            self.send_header('Content-Length', str(len(excel_data)))
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(excel_data)
            
        except Exception as e:
            error_response = {
                "success": False,
                "error": f"Erro ao gerar arquivo: {str(e)}"
            }
            self.send_response(500)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(error_response, ensure_ascii=False).encode('utf-8'))
    
    def create_excel_data(self, resultados):
        """Cria dados Excel a partir dos resultados"""
        try:
            # Criar workbook e worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Resultados Análise Coletiva"
            
            # Definir cabeçalhos
            headers = ['Idade', 'Sexo', 'Período', 'Tipo Tábua', 'Tábua', 'Taxa à Vista', 'Taxa Mensal']
            
            # Escrever cabeçalhos (sem formatação complexa por enquanto)
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Escrever dados
            for row, resultado in enumerate(resultados, 2):
                sexo_texto = 'Masculino' if resultado['sexo'] == 'M' else 'Feminino'
                
                ws.cell(row=row, column=1, value=resultado['idade'])
                ws.cell(row=row, column=2, value=sexo_texto)
                ws.cell(row=row, column=3, value=resultado['periodo'])
                ws.cell(row=row, column=4, value=resultado['tipo_tabua'])
                ws.cell(row=row, column=5, value=resultado['tabua'])
                ws.cell(row=row, column=6, value=resultado['taxa_vista'])
                ws.cell(row=row, column=7, value=resultado['taxa_mensal'])
            
            # Ajustar largura das colunas
            column_widths = [8, 12, 10, 12, 15, 15, 15]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
            
            # Salvar em BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Retornar os dados como bytes
            excel_bytes = output.getvalue()
            output.close()
            
            return excel_bytes
            
        except Exception as e:
            print(f"Erro ao criar Excel: {e}")
            # Fallback: criar um arquivo simples
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Resultados"
            
            # Cabeçalhos simples
            headers = ['Idade', 'Sexo', 'Período', 'Tipo Tábua', 'Tábua', 'Taxa à Vista', 'Taxa Mensal']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Dados simples
            for row, resultado in enumerate(resultados, 2):
                sexo_texto = 'Masculino' if resultado['sexo'] == 'M' else 'Feminino'
                ws.cell(row=row, column=1, value=resultado['idade'])
                ws.cell(row=row, column=2, value=sexo_texto)
                ws.cell(row=row, column=3, value=resultado['periodo'])
                ws.cell(row=row, column=4, value=resultado['tipo_tabua'])
                ws.cell(row=row, column=5, value=resultado['tabua'])
                ws.cell(row=row, column=6, value=resultado['taxa_vista'])
                ws.cell(row=row, column=7, value=resultado['taxa_mensal'])
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            excel_bytes = output.getvalue()
            output.close()
            
            return excel_bytes
    
    def handle_cache_stats(self):
        """Retorna estatísticas do cache."""
        try:
            stats = obter_estatisticas_cache()
            
            self.send_response(200)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(stats, ensure_ascii=False).encode('utf-8'))
            
        except Exception as e:
            error_response = {"success": False, "error": str(e)}
            self.send_response(500)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(error_response, ensure_ascii=False).encode('utf-8'))
    
    def handle_limpar_cache(self):
        """Limpa o cache de tábuas."""
        try:
            limpar_cache_tabuas()
            
            response = {"success": True, "message": "Cache limpo com sucesso"}
            
            self.send_response(200)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(response, ensure_ascii=False).encode('utf-8'))
            
        except Exception as e:
            error_response = {"success": False, "error": str(e)}
            self.send_response(500)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(error_response, ensure_ascii=False).encode('utf-8'))
    
    def handle_calcular_coletivo_postalis(self):
        """Calcula seguro prestamista coletivo."""
        try:
            # Ler dados do POST
            content_length = int(self.headers.get('Content-Length', 0))
            if content_length == 0:
                raise ValueError("Content-Length is 0")
            
            post_data = self.rfile.read(content_length)
            if not post_data:
                raise ValueError("No data received")
            
            data = json.loads(post_data.decode('utf-8'))
            
            # Extrair parâmetros
            idade_min = int(data['idade_min'])
            idade_max = int(data['idade_max'])
            sexos = data['sexos']
            parcelas_min = int(data['parcelas_min'])
            parcelas_max = int(data['parcelas_max'])
            taxa_juros = float(data['taxa_juros']) / 100
            valor_financiamento = float(data['valor_financiamento'])
            periodo_total = int(data['periodo_total'])
            tabuas_validas = data['tabuas_validas']
            tabuas_invalidas = data['tabuas_invalidas']
            
            # Validação dos limites
            if not (0 <= idade_min <= idade_max <= 99):
                raise ValueError("As idades devem estar entre 0 e 99 anos, e a idade mínima deve ser menor ou igual à máxima.")
            
            if not (1 <= parcelas_min <= parcelas_max <= 120):
                raise ValueError("As parcelas restantes devem estar entre 1 e 120, e as parcelas mínimas devem ser menores ou iguais às máximas.")
            
            if not (0 <= taxa_juros <= 0.30):  # 0% a 30%
                raise ValueError("A taxa de juros deve estar entre 0% e 30%.")
            
            if not (1 <= valor_financiamento <= 500000):
                raise ValueError("O valor do financiamento deve estar entre R$ 1,00 e R$ 500.000,00.")
            
            if not (periodo_total == 120):
                raise ValueError("O período total deve ser 120 meses.")
            
            if len(sexos) == 0:
                raise ValueError("Selecione pelo menos um sexo.")
            
            if len(tabuas_validas) == 0 and len(tabuas_invalidas) == 0:
                raise ValueError("Selecione pelo menos uma tábua válida ou inválida.")
            
            # Calcular total de combinações
            total_idades = idade_max - idade_min + 1
            total_parcelas = parcelas_max - parcelas_min + 1
            total_sexos = len(sexos)
            total_tabuas = len(tabuas_validas) + len(tabuas_invalidas)
            total_combinacoes = total_idades * total_parcelas * total_sexos * total_tabuas
            
            print(f"Iniciando calculo coletivo prestamista:")
            print(f"   • Idades: {idade_min}-{idade_max} ({total_idades} idades)")
            print(f"   • Parcelas: {parcelas_min}-{parcelas_max} ({total_parcelas} parcelas)")
            print(f"   • Sexos: {sexos} ({total_sexos} sexos)")
            print(f"   • Tábuas: {len(tabuas_validas)} válidas + {len(tabuas_invalidas)} inválidas")
            print(f"   • Total: {total_combinacoes} combinações")
            
            # Processar combinações
            resultados = []
            tabuas_utilizadas = set()
            
            # Processar tábuas válidas
            for tabua in tabuas_validas:
                tabuas_utilizadas.add(tabua)
                for idade in range(idade_min, idade_max + 1):
                    for sexo in sexos:
                        for parcelas_restantes in range(parcelas_min, parcelas_max + 1):
                            try:
                                # Calcular seguro prestamista
                                tabua_obj = TabuladeComutacao(taxa_juros, tabua)
                                tabua_obj.tabua_selecionada = tabua
                                
                                if tabua not in tabua_obj.tabuas_disponiveis:
                                    # Se a tábua não for encontrada, usar AT-83 como padrão
                                    if 'AT-83' in tabua_obj.tabuas_disponiveis:
                                        tabua = 'AT-83'
                                        tabua_obj.tabua_selecionada = tabua
                                    elif tabua_obj.tabuas_disponiveis:
                                        tabua = list(tabua_obj.tabuas_disponiveis.keys())[0]
                                        tabua_obj.tabua_selecionada = tabua
                                    else:
                                        continue
                                
                                # Calcular seguro prestamista
                                resultado_prestamista = calcular_seguro_prestamista(
                                    tabua_obj, idade, sexo, periodo_total, taxa_juros, valor_financiamento
                                )
                                
                                # Calcular saldo devedor na parcela atual
                                parcela_atual = periodo_total - parcelas_restantes + 1
                                taxa_mensal = (1 + taxa_juros)**(1/12) - 1
                                saldo_devedor_atual = calcular_saldo_devedor_price(
                                    valor_financiamento, taxa_mensal, periodo_total, parcela_atual
                                )
                                
                                # Calcular taxas de risco
                                taxa_risco_anual = (resultado_prestamista['premio_unico'] / saldo_devedor_atual) * 100 if saldo_devedor_atual > 0 else 0
                                taxa_risco_mensal = (resultado_prestamista['premio_mensal'] / saldo_devedor_atual) * 100 if saldo_devedor_atual > 0 else 0
                                
                                resultado = {
                                    "idade": idade,
                                    "sexo": sexo,
                                    "parcelas_restantes": parcelas_restantes,
                                    "tipo_tabua": "Válido",
                                    "tabua": tabua,
                                    "premio_anual": resultado_prestamista['premio_unico'],
                                    "premio_mensal": resultado_prestamista['premio_mensal'],
                                    "taxa_risco_anual": taxa_risco_anual,
                                    "taxa_risco_mensal": taxa_risco_mensal
                                }
                                resultados.append(resultado)
                                
                            except Exception as e:
                                print(f"Erro ao calcular {idade}, {sexo}, {parcelas_restantes}, {tabua}: {e}")
                                pass
            
            # Processar tábuas inválidas
            for tabua in tabuas_invalidas:
                tabuas_utilizadas.add(tabua)
                for idade in range(idade_min, idade_max + 1):
                    for sexo in sexos:
                        for parcelas_restantes in range(parcelas_min, parcelas_max + 1):
                            try:
                                # Calcular seguro prestamista
                                tabua_obj = TabuladeComutacao(taxa_juros, tabua)
                                tabua_obj.tabua_selecionada = tabua
                                
                                if tabua not in tabua_obj.tabuas_disponiveis:
                                    # Se a tábua não for encontrada, usar AT-83 como padrão
                                    if 'AT-83' in tabua_obj.tabuas_disponiveis:
                                        tabua = 'AT-83'
                                        tabua_obj.tabua_selecionada = tabua
                                    elif tabua_obj.tabuas_disponiveis:
                                        tabua = list(tabua_obj.tabuas_disponiveis.keys())[0]
                                        tabua_obj.tabua_selecionada = tabua
                                    else:
                                        continue
                                
                                # Calcular seguro prestamista
                                resultado_prestamista = calcular_seguro_prestamista(
                                    tabua_obj, idade, sexo, periodo_total, taxa_juros, valor_financiamento
                                )
                                
                                # Calcular saldo devedor na parcela atual
                                parcela_atual = periodo_total - parcelas_restantes + 1
                                taxa_mensal = (1 + taxa_juros)**(1/12) - 1
                                saldo_devedor_atual = calcular_saldo_devedor_price(
                                    valor_financiamento, taxa_mensal, periodo_total, parcela_atual
                                )
                                
                                # Calcular taxas de risco
                                taxa_risco_anual = (resultado_prestamista['premio_unico'] / saldo_devedor_atual) * 100 if saldo_devedor_atual > 0 else 0
                                taxa_risco_mensal = (resultado_prestamista['premio_mensal'] / saldo_devedor_atual) * 100 if saldo_devedor_atual > 0 else 0
                                
                                resultado = {
                                    "idade": idade,
                                    "sexo": sexo,
                                    "parcelas_restantes": parcelas_restantes,
                                    "tipo_tabua": "Inválido",
                                    "tabua": tabua,
                                    "premio_anual": resultado_prestamista['premio_unico'],
                                    "premio_mensal": resultado_prestamista['premio_mensal'],
                                    "taxa_risco_anual": taxa_risco_anual,
                                    "taxa_risco_mensal": taxa_risco_mensal
                                }
                                resultados.append(resultado)
                                
                            except Exception as e:
                                print(f"Erro ao calcular {idade}, {sexo}, {parcelas_restantes}, {tabua}: {e}")
                                pass
            
            # Preparar resposta
            response = {
                "success": True,
                "resultados": resultados,
                "total_combinacoes": total_combinacoes,
                "combinacoes_processadas": len(resultados),
                "progresso": 100.0,
                "idade_min": idade_min,
                "idade_max": idade_max,
                "parcelas_min": parcelas_min,
                "parcelas_max": parcelas_max,
                "tabuas_utilizadas": list(tabuas_utilizadas)
            }
            
            # Enviar resposta
            self.send_response(200)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(response, ensure_ascii=False).encode('utf-8'))
            
        except Exception as e:
            error_response = {"success": False, "error": str(e)}
            self.send_response(500)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(error_response, ensure_ascii=False).encode('utf-8'))
    
    def handle_download_excel_postalis(self):
        """Gera arquivo Excel para resultados do seguro prestamista coletivo."""
        try:
            # Ler dados do POST
            content_length = int(self.headers.get('Content-Length', 0))
            if content_length == 0:
                raise ValueError("Content-Length is 0")
            
            post_data = self.rfile.read(content_length)
            if not post_data:
                raise ValueError("No data received")
            
            data = json.loads(post_data.decode('utf-8'))
            resultados = data.get('resultados', [])
            
            if not resultados:
                raise ValueError("Nenhum resultado para exportar")
            
            # Criar arquivo Excel
            excel_data = self.create_excel_data_postalis(resultados)
            
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', 'attachment; filename="resultados_seguro_prestamista_coletivo.xlsx"')
            self.send_header('Content-Length', str(len(excel_data)))
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(excel_data)
            
        except Exception as e:
            error_response = {
                "success": False,
                "error": f"Erro ao gerar arquivo: {str(e)}"
            }
            self.send_response(500)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(error_response, ensure_ascii=False).encode('utf-8'))
    
    def create_excel_data_postalis(self, resultados):
        """Cria dados Excel para resultados do seguro prestamista coletivo."""
        try:
            # Criar workbook e worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Resultados Seguro Prestamista Coletivo"
            
            # Definir cabeçalhos
            headers = ['Idade', 'Sexo', 'Parcelas Restantes', 'Tipo Tábua', 'Tábua', 'Prêmio Anual (R$)', 'Prêmio Mensal (R$)', 'Taxa Risco Anual (%)', 'Taxa Risco Mensal (%)']
            
            # Escrever cabeçalhos
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Escrever dados
            for row, resultado in enumerate(resultados, 2):
                sexo_texto = 'Masculino' if resultado['sexo'] == 'M' else 'Feminino'
                
                ws.cell(row=row, column=1, value=resultado['idade'])
                ws.cell(row=row, column=2, value=sexo_texto)
                ws.cell(row=row, column=3, value=resultado['parcelas_restantes'])
                ws.cell(row=row, column=4, value=resultado['tipo_tabua'])
                ws.cell(row=row, column=5, value=resultado['tabua'])
                ws.cell(row=row, column=6, value=resultado['premio_anual'])
                ws.cell(row=row, column=7, value=resultado['premio_mensal'])
                ws.cell(row=row, column=8, value=resultado['taxa_risco_anual'])
                ws.cell(row=row, column=9, value=resultado['taxa_risco_mensal'])
            
            # Ajustar largura das colunas
            column_widths = [8, 12, 15, 12, 15, 18, 18, 18, 18]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
            
            # Salvar em BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Retornar os dados como bytes
            excel_bytes = output.getvalue()
            output.close()
            
            return excel_bytes
            
        except Exception as e:
            print(f"Erro ao criar Excel: {e}")
            # Fallback: criar um arquivo simples
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Resultados"
            
            # Cabeçalhos simples
            headers = ['Idade', 'Sexo', 'Parcelas Restantes', 'Tipo Tábua', 'Tábua', 'Prêmio Anual (R$)', 'Prêmio Mensal (R$)', 'Taxa Risco Anual (%)', 'Taxa Risco Mensal (%)']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Dados simples
            for row, resultado in enumerate(resultados, 2):
                sexo_texto = 'Masculino' if resultado['sexo'] == 'M' else 'Feminino'
                ws.cell(row=row, column=1, value=resultado['idade'])
                ws.cell(row=row, column=2, value=sexo_texto)
                ws.cell(row=row, column=3, value=resultado['parcelas_restantes'])
                ws.cell(row=row, column=4, value=resultado['tipo_tabua'])
                ws.cell(row=row, column=5, value=resultado['tabua'])
                ws.cell(row=row, column=6, value=resultado['premio_anual'])
                ws.cell(row=row, column=7, value=resultado['premio_mensal'])
                ws.cell(row=row, column=8, value=resultado['taxa_risco_anual'])
                ws.cell(row=row, column=9, value=resultado['taxa_risco_mensal'])
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            excel_bytes = output.getvalue()
            output.close()
            
            return excel_bytes
    
    def handle_calcular_prestamista(self):
        """Calcula seguro prestamista."""
        try:
            # Ler dados do POST
            content_length = int(self.headers.get('Content-Length', 0))
            if content_length == 0:
                raise ValueError("Content-Length is 0")
            
            post_data = self.rfile.read(content_length)
            if not post_data:
                raise ValueError("No data received")
            
            data = json.loads(post_data.decode('utf-8'))
            
            # Extrair parâmetros
            idade = int(data['idade'])
            sexo = data['sexo']
            periodo = int(data['periodo'])
            taxa_juros = float(data['taxa_juros']) / 100
            soma_segurada = float(data['soma_segurada'])
            tabua_selecionada = data['tabua_mortalidade']
            
            # Validação dos limites
            if not (0 <= idade <= 110):
                raise ValueError("A idade deve estar entre 0 e 110 anos.")
            
            if not (1 <= periodo <= 120):
                raise ValueError("O período deve estar entre 1 e 120 meses.")
            
            if not (0 <= taxa_juros <= 0.20):  # 0% a 20%
                raise ValueError("A taxa de juros deve estar entre 0% e 20%.")
            
            if not (0 <= soma_segurada <= 200000):
                raise ValueError("A soma segurada deve estar entre R$ 0,00 e R$ 200.000,00.")
            
            # Calcular seguro prestamista
            tabua_obj = TabuladeComutacao(taxa_juros, tabua_selecionada)
            tabua_obj.tabua_selecionada = tabua_selecionada
            
            # Verificar se a tábua existe
            if tabua_selecionada not in tabua_obj.tabuas_disponiveis:
                raise KeyError(f"Tábua '{tabua_selecionada}' não encontrada. Tábuas disponíveis: {list(tabua_obj.tabuas_disponiveis.keys())}")
            
            resultado_prestamista = calcular_seguro_prestamista(tabua_obj, idade, sexo, periodo, taxa_juros, soma_segurada)
            
            # Preparar resposta
            response = {
                "success": True,
                "tipo_seguro": "Prestamista",
                "valor_total": f"R$ {resultado_prestamista['premio_unico']:,.2f}",
                "percentual_total": f"{resultado_prestamista['premio_unico']/soma_segurada*100:.4f}%",
                "valor_mensal": f"R$ {resultado_prestamista['premio_mensal']:,.2f}",
                "percentual_mensal": f"{resultado_prestamista['percentual_mensal']*100:.4f}%",
                "taxa_quitação_risco_mensal": f"{resultado_prestamista['taxa_quitação_risco_mensal']*100:.4f}%",
                "detalhes": {
                    "dados_entrada": {
                        "idade": idade,
                        "sexo": "Masculino" if sexo == 'M' else "Feminino",
                        "periodo": periodo,
                        "taxa_juros": f"{taxa_juros*100:.4f}%",
                        "soma_segurada": f"R$ {soma_segurada:,.2f}",
                        "tabua": tabua_selecionada
                    },
                    "financiamento": {
                        "taxa_mensal": f"{resultado_prestamista['taxa_mensal']*100:.4f}%",
                        "num_parcelas": resultado_prestamista['num_parcelas'],
                        "pmt_financiamento": f"R$ {resultado_prestamista['pmt_financiamento']:,.2f}",
                        "formula_saldo_devedor": r"SD(t) = S_0 \times \frac{(1+i)^n - (1+i)^t}{(1+i)^n - 1}"
                    },
                    "calculo_premio": {
                        "formula_premio_unico": r"A = \sum_{t=0}^{n-1} SD(t) \times q_{x+t} \times v^{t+1}",
                        "premio_unico": f"R$ {resultado_prestamista['premio_unico']:,.2f}",
                        "premio_mensal": f"R$ {resultado_prestamista['premio_mensal']:,.2f}",
                        "percentual_mensal": f"{resultado_prestamista['percentual_mensal']*100:.4f}%",
                        "taxa_quitação_risco_mensal": f"{resultado_prestamista['taxa_quitação_risco_mensal']*100:.4f}%"
                    },
                    "evolucao_saldo": resultado_prestamista['detalhes_calculo']  # Todos os meses
                }
            }
            
            # Enviar resposta
            self.send_response(200)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(response, ensure_ascii=False).encode('utf-8'))
            
        except Exception as e:
            error_response = {"success": False, "error": str(e)}
            self.send_response(500)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(error_response, ensure_ascii=False).encode('utf-8'))
    
    def handle_calcular_reserva_matematica(self):
        """Calcula a reserva matemática do seguro prestamista no tempo t."""
        try:
            # Ler dados do POST
            content_length = int(self.headers.get('Content-Length', 0))
            if content_length == 0:
                raise ValueError("Content-Length is 0")
            
            post_data = self.rfile.read(content_length)
            if not post_data:
                raise ValueError("No data received")
            
            data = json.loads(post_data.decode('utf-8'))
            
            # Extrair parâmetros
            idade = int(data['idade'])
            sexo = data['sexo']
            periodo = int(data['periodo'])
            taxa_juros = float(data['taxa_juros']) / 100
            soma_segurada = float(data['soma_segurada'])
            tabua_selecionada = data['tabua_mortalidade']
            tempo_t = int(data['tempo_t'])
            
            # Validação dos limites
            if not (0 <= idade <= 110):
                raise ValueError("A idade deve estar entre 0 e 110 anos.")
            
            if not (1 <= periodo <= 120):
                raise ValueError("O período deve estar entre 1 e 120 meses.")
            
            if not (0 <= taxa_juros <= 0.20):  # 0% a 20%
                raise ValueError("A taxa de juros deve estar entre 0% e 20%.")
            
            if not (0 <= soma_segurada <= 200000):
                raise ValueError("A soma segurada deve estar entre R$ 0,00 e R$ 200.000,00.")
            
            if not (0 <= tempo_t <= periodo):
                raise ValueError(f"O tempo t deve estar entre 0 e {periodo} meses.")
            
            # Calcular reserva matemática
            tabua_obj = TabuladeComutacao(taxa_juros, tabua_selecionada)
            tabua_obj.tabua_selecionada = tabua_selecionada
            
            # Verificar se a tábua existe
            if tabua_selecionada not in tabua_obj.tabuas_disponiveis:
                raise KeyError(f"Tábua '{tabua_selecionada}' não encontrada. Tábuas disponíveis: {list(tabua_obj.tabuas_disponiveis.keys())}")
            
            # Calcular o prêmio mensal do seguro prestamista primeiro
            resultado_prestamista = calcular_seguro_prestamista(tabua_obj, idade, sexo, periodo, taxa_juros, soma_segurada)
            premio_mensal = resultado_prestamista['premio_mensal']
            
            resultado_reserva = calcular_reserva_matematica_prestamista(
                tabua_obj, idade, sexo, periodo, taxa_juros, soma_segurada, tempo_t, premio_mensal
            )
            
            # Preparar resposta
            response = {
                "success": True,
                "tipo_calculo": "Reserva Matemática",
                "reserva_matematica": f"R$ {resultado_reserva['reserva_matematica']:,.2f}",
                "valor_presente_beneficios": f"R$ {resultado_reserva['valor_presente_beneficios']:,.2f}",
                "valor_presente_premios": f"R$ {resultado_reserva['valor_presente_premios']:,.2f}",
                "premio_mensal": f"R$ {premio_mensal:,.2f}",
                "explicacao": resultado_reserva['explicacao'],
                "detalhes": {
                    "dados_entrada": {
                        "idade_inicial": idade,
                        "sexo": "Masculino" if sexo == 'M' else "Feminino",
                        "periodo_total": periodo,
                        "taxa_juros": f"{taxa_juros*100:.4f}%",
                        "soma_segurada": f"R$ {soma_segurada:,.2f}",
                        "tabua": tabua_selecionada,
                        "tempo_t": tempo_t
                    },
                    "situacao_atual": {
                        "idade_atual": resultado_reserva['detalhes']['idade_atual'],
                        "periodo_restante": resultado_reserva['detalhes']['periodo_restante'],
                        "saldo_devedor_atual": f"R$ {resultado_reserva['detalhes']['saldo_devedor_atual']:,.2f}",
                        "valor_segurado_atual": f"R$ {resultado_reserva['detalhes']['valor_segurado_atual']:,.2f}",
                        "prob_sobrevivencia_t": f"{resultado_reserva['detalhes']['prob_sobrevivencia_t']:.6f}"
                    },
                    "formula_reserva": {
                        "formula_geral": r"V_t = \sum_{k=t+1}^{n} v^{k-t} B_k \cdot _{k-1-t}P_{x+t} \cdot q_{x+k-1} - \sum_{k=t+1}^{n} v^{k-t} P_k \cdot _{k-t}P_{x+t}",
                        "explicacao": "Reserva matemática prospectiva no tempo t usando a fórmula padrão atuarial.",
                        "simbolos": {
                            "V_t": "Reserva matemática no tempo t",
                            "B_k": "Benefício no mês k (saldo devedor)",
                            "P_k": "Prêmio no mês k (prêmio mensal do seguro)",
                            "_{k-1-t}P_{x+t}": "Probabilidade de sobrevivência até k-1, condicional em ter sobrevivido até t",
                            "_{k-t}P_{x+t}": "Probabilidade de sobrevivência até k, condicional em ter sobrevivido até t",
                            "q_{x+k-1}": "Probabilidade de morte no mês k",
                            "v^{k-t}": "Fator de desconto de t para k"
                        }
                    },
                    "calculo_detalhado": resultado_reserva['detalhes']['calculo_detalhado']
                }
            }
            
            # Enviar resposta
            self.send_response(200)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(response, ensure_ascii=False).encode('utf-8'))
            
        except Exception as e:
            error_response = {"success": False, "error": str(e)}
            self.send_response(500)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(error_response, ensure_ascii=False).encode('utf-8'))
    
    def handle_calcular_prestamista_alt(self):
        """Calcula seguro prestamista com metodologia alternativa."""
        try:
            # Ler dados do POST
            content_length = int(self.headers.get('Content-Length', 0))
            if content_length == 0:
                raise ValueError("Content-Length is 0")
            
            post_data = self.rfile.read(content_length)
            if not post_data:
                raise ValueError("No data received")
            
            data = json.loads(post_data.decode('utf-8'))
            
            # Extrair parâmetros
            idade = int(data['idade'])
            sexo = data['sexo']
            periodo = int(data['periodo'])
            taxa_juros = float(data['taxa_juros']) / 100
            soma_segurada = float(data['soma_segurada'])
            tabua_selecionada = data['tabua_mortalidade']
            
            # Validação dos limites
            if not (0 <= idade <= 110):
                raise ValueError("A idade deve estar entre 0 e 110 anos.")
            
            if not (1 <= periodo <= 120):
                raise ValueError("O período deve estar entre 1 e 120 meses.")
            
            if not (0 <= taxa_juros <= 0.20):  # 0% a 20%
                raise ValueError("A taxa de juros deve estar entre 0% e 20%.")
            
            if not (0 <= soma_segurada <= 200000):
                raise ValueError("A soma segurada deve estar entre R$ 0,00 e R$ 200.000,00.")
            
            # Calcular seguro prestamista com metodologia alternativa
            tabua_obj = TabuladeComutacao(taxa_juros, tabua_selecionada)
            tabua_obj.tabua_selecionada = tabua_selecionada
            
            # Verificar se a tábua existe
            if tabua_selecionada not in tabua_obj.tabuas_disponiveis:
                raise KeyError(f"Tábua '{tabua_selecionada}' não encontrada. Tábuas disponíveis: {list(tabua_obj.tabuas_disponiveis.keys())}")
            
            # Calcular seguro prestamista com metodologia alternativa
            resultado_prestamista = calcular_seguro_prestamista_alt(tabua_obj, idade, sexo, periodo, taxa_juros, soma_segurada)
            
            # Preparar resposta
            response = {
                "success": True,
                "tipo_seguro": "Prestamista - Metodologia Alternativa",
                "valor_total": f"R$ {resultado_prestamista['premio_unico']:,.2f}",
                "percentual_total": f"{resultado_prestamista['premio_unico']/soma_segurada*100:.4f}%",
                "valor_mensal": f"R$ {resultado_prestamista['premio_mensal']:,.2f}",
                "percentual_mensal": f"{resultado_prestamista['percentual_mensal']*100:.4f}%",
                "taxa_cobertura_risco": f"{resultado_prestamista['taxa_cobertura_risco']*100:.4f}%",
                "valor_emprestimo_atual": f"R$ {resultado_prestamista['valor_emprestimo_atual']:,.2f}",
                "detalhes": {
                    "dados_entrada": {
                        "idade": idade,
                        "sexo": "Masculino" if sexo == 'M' else "Feminino",
                        "periodo": periodo,
                        "taxa_juros": f"{taxa_juros*100:.4f}%",
                        "soma_segurada": f"R$ {soma_segurada:,.2f}",
                        "tabua": tabua_selecionada
                    },
                    "financiamento": {
                        "taxa_mensal": f"{resultado_prestamista['taxa_mensal']*100:.4f}%",
                        "num_parcelas": resultado_prestamista['num_parcelas'],
                        "pmt_financiamento": f"R$ {resultado_prestamista['pmt_financiamento']:,.2f}",
                        "formula_saldo_devedor": r"SD(t) = S_0 \times \frac{(1+i)^n - (1+i)^t}{(1+i)^n - 1}"
                    },
                    "calculo_premio": {
                        "formula_premio_unico": r"A = \sum_{t=0}^{n-1} SD(t) \times q_{x+t} \times v^{t+1}",
                        "premio_unico": f"R$ {resultado_prestamista['premio_unico']:,.2f}",
                        "premio_mensal": f"R$ {resultado_prestamista['premio_mensal']:,.2f}",
                        "percentual_mensal": f"{resultado_prestamista['percentual_mensal']*100:.4f}%",
                        "taxa_cobertura_risco": f"{resultado_prestamista['taxa_cobertura_risco']*100:.4f}%",
                        "soma_vp_pgto": f"R$ {resultado_prestamista['soma_vp_pgto']:,.2f}",
                        "soma_vpa_pgto": f"R$ {resultado_prestamista['soma_vpa_pgto']:,.2f}",
                        "valor_emprestimo_atual": f"R$ {resultado_prestamista['valor_emprestimo_atual']:,.2f}"
                    },
                    "evolucao_saldo": resultado_prestamista['detalhes_calculo']  # Todos os meses
                }
            }
            
            # Enviar resposta
            self.send_response(200)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(response, ensure_ascii=False).encode('utf-8'))
            
        except Exception as e:
            error_response = {"success": False, "error": str(e)}
            self.send_response(500)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(error_response, ensure_ascii=False).encode('utf-8'))
    
    def do_POST(self):
        if self.path == '/calcular':
            try:
                # Ler dados do POST
                content_length = int(self.headers.get('Content-Length', 0))
                if content_length == 0:
                    raise ValueError("Content-Length is 0")
                
                post_data = self.rfile.read(content_length)
                if not post_data:
                    raise ValueError("No data received")
                
                data = json.loads(post_data.decode('utf-8'))
                
                # Extrair parâmetros
                idade = int(data['idade'])
                sexo = data['sexo']
                periodo = int(data['periodo'])
                taxa_juros = float(data['taxa_juros']) / 100
                soma_segurada = float(data['soma_segurada'])
                tabua_selecionada = data['tabua_mortalidade']
                
                # Validação dos limites
                if not (0 <= idade <= 110):
                    raise ValueError("A idade deve estar entre 0 e 110 anos.")
                
                if not (1 <= periodo <= 120):
                    raise ValueError("O período deve estar entre 1 e 120 meses.")
                
                if not (0 <= taxa_juros <= 0.20):  # 0% a 20%
                    raise ValueError("A taxa de juros deve estar entre 0% e 20%.")
                
                if not (0 <= soma_segurada <= 200000):
                    raise ValueError("A soma segurada deve estar entre R$ 0,00 e R$ 200.000,00.")
                
                fracionamento = 12  # Mensal
                
                # Calcular
                tabua_obj = TabuladeComutacao(taxa_juros, tabua_selecionada)
                tabua_obj.tabua_selecionada = tabua_selecionada
                
                # Verificar se a tábua existe
                if tabua_selecionada not in tabua_obj.tabuas_disponiveis:
                    raise KeyError(f"Tábua '{tabua_selecionada}' não encontrada. Tábuas disponíveis: {list(tabua_obj.tabuas_disponiveis.keys())}")
                
                tabua_obj.dados = tabua_obj.tabuas_disponiveis[tabua_selecionada]
                tabua_obj.calcular_tabua_comutacao()
                
                seguro_anual = calcular_seguro_anual(tabua_obj, idade, sexo, periodo)
                seguro_fracionado_total, fator_ajuste, taxa_fracionada_calc = calcular_seguro_fracionado_total(taxa_juros, fracionamento, seguro_anual)
                valor_monetario_vista = soma_segurada * seguro_fracionado_total
                
                valor_mensal, N_x, N_x_n, anuidade_ajustada, percentual_mensal_calc = calcular_premio_mensal(tabua_obj, idade, periodo, valor_monetario_vista, taxa_juros, fracionamento, soma_segurada)
                
                # Preparar resposta
                response = {
                    "success": True,
                    "valor_total": f"R$ {valor_monetario_vista:,.2f}",
                    "percentual_total": f"{seguro_fracionado_total*100:.4f}%",
                    "valor_mensal": f"R$ {valor_mensal:,.2f}",
                    "percentual_mensal": f"{percentual_mensal_calc[0]*100:.4f}%",
                    "detalhes": {
                        "dados_entrada": {
                            "idade": idade,
                            "sexo": "Masculino" if sexo == 'M' else "Feminino",
                            "periodo": periodo,
                            "taxa_juros": f"{taxa_juros*100:.4f}%",
                            "fracionamento": "12 vezes por ano (mensal)",
                            "soma_segurada": f"R$ {soma_segurada:,.2f}"
                        },
                        "calculo_seguro_anual": {
                            "M_x": f"{tabua_obj.M_x[idade]:.2f}",
                            "M_x_n": f"{tabua_obj.M_x[idade + periodo]:.2f}",
                            "D_x": f"{tabua_obj.D_x[idade]:.2f}",
                            "formula": r"A_{x:n}^1 = \frac{M_x - M_{x+n}}{D_x}",
                            "valor": f"{seguro_anual:.8f}"
                        },
                        "calculo_seguro_fracionado": {
                            "taxa_fracionada_formula": r"\text{Taxa fracionada} = k \times ((1+i)^{1/k} - 1)",
                            "taxa_fracionada_valor": f"{taxa_fracionada_calc:.6f} ({taxa_fracionada_calc*100:.4f}%)",
                            "fator_ajuste_formula": r"\text{Fator de ajuste} = \frac{i}{k \times ((1+i)^{1/k} - 1)}",
                            "fator_ajuste_valor": f"{fator_ajuste:.6f}",
                            "formula": r"A_{x:n}^{(k)} = \left( \frac{i}{k \times ((1+i)^{1/k} - 1)} \right) \times A_{x:n}^1",
                            "valor": f"{seguro_fracionado_total:.8f}"
                        },
                        "calculo_premio_vista": {
                            "valor_unitario": f"{seguro_fracionado_total:.8f}",
                            "valor_monetario": f"R$ {valor_monetario_vista:,.2f}"
                        },
                        "calculo_premio_mensal": {
                            "N_x": f"{N_x:.2f}",
                            "N_x_n": f"{N_x_n:.2f}",
                            "D_x": f"{tabua_obj.D_x[idade]:.2f}",
                            "D_x_n": f"{tabua_obj.D_x[idade + periodo]:.2f}",
                            "anuidade_ajustada_formula": r"\text{Anuidade Ajustada} = \left( \frac{N_x - N_{x+n}}{D_x} + \frac{11}{24} \times \left(1 - \frac{D_{x+n}}{D_x}\right) \right) \times k",
                            "anuidade_ajustada_valor": f"{anuidade_ajustada:.6f}",
                            "valor_mensal_formula": r"\text{Valor Mensal} = \frac{\text{Valor Monetário (à vista)}}{\text{Anuidade Ajustada}}",
                            "valor_mensal_valor": f"R$ {valor_mensal:,.2f}",
                            "percentual_mensal_formula": r"\text{Percentual Mensal} = \frac{\text{Valor Mensal}}{\text{PGTO}(\text{taxa_fracionada}, \text{num_pagamentos}, -\text{Soma Segurada})}",
                            "percentual_mensal_valor": f"{percentual_mensal_calc[0]*100:.4f}%"
                        },
                        "tabua_comutacao": {
                            "l_x": {str(k): f"{v:.2f}" for k, v in tabua_obj.l_x.items()},
                            "d_x": {str(k): f"{v:.2f}" for k, v in tabua_obj.d_x.items()},
                            "D_x": {str(k): f"{v:.2f}" for k, v in tabua_obj.D_x.items()},
                            "C_x": {str(k): f"{v:.2f}" for k, v in tabua_obj.C_x.items()},
                            "N_x": {str(k): f"{v:.2f}" for k, v in tabua_obj.N_x.items()},
                            "M_x": {str(k): f"{v:.2f}" for k, v in tabua_obj.M_x.items()},
                            "v_x": {str(k): f"{v:.6f}" for k, v in tabua_obj.v_x.items()}
                        }
                    }
                }
                
                # Enviar resposta
                self.send_response(200)
                self.send_header('Content-type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps(response, ensure_ascii=False).encode('utf-8'))
                
            except Exception as e:
                error_response = {"success": False, "error": str(e)}
                self.send_response(500)
                self.send_header('Content-type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps(error_response, ensure_ascii=False).encode('utf-8'))
        elif self.path == '/calcular_coletivo':
            self.handle_calcular_coletivo()
        elif self.path == '/calcular_coletivo_progress':
            self.handle_calcular_coletivo_progress()
        elif self.path == '/calcular_coletivo_postalis':
            self.handle_calcular_coletivo_postalis()
        elif self.path == '/download_excel':
            self.handle_download_excel()
        elif self.path == '/download_excel_postalis':
            self.handle_download_excel_postalis()
        elif self.path == '/calcular_prestamista':
            self.handle_calcular_prestamista()
        elif self.path == '/calcular_prestamista_alt':
            self.handle_calcular_prestamista_alt()
        elif self.path == '/calcular_reserva_matematica':
            self.handle_calcular_reserva_matematica()
        elif self.path == '/calcular_reserva_matematica_individual':
            self.handle_calcular_reserva_matematica_individual()
        elif self.path == '/calcular_reserva_matematica_coletiva':
            self.handle_calcular_reserva_matematica_coletiva()
        elif self.path == '/preview_planilha':
            self.handle_preview_planilha()
        elif self.path == '/obter_qx':
            self.handle_obter_qx()
        elif self.path == '/obter_tabua_completa':
            self.handle_obter_tabua_completa()
        else:
            self.send_response(404)
            self.end_headers()

    def handle_calcular_reserva_matematica_coletiva(self):
        try:
            import pandas as pd
            import io
            
            # Ler dados do POST
            content_length = int(self.headers.get('Content-Length', 0))
            if content_length == 0:
                raise ValueError("Content-Length is 0")
            
            post_data = self.rfile.read(content_length)
            if not post_data:
                raise ValueError("No data received")
            
            # Parse multipart/form-data
            content_type = self.headers.get('Content-Type', '')
            if 'boundary=' not in content_type:
                raise ValueError("Boundary não encontrado no Content-Type")
            
            boundary = content_type.split('boundary=')[1]
            parts = post_data.split(f'--{boundary}'.encode())
            
            form_data = {}
            taxas_file_data = None
            emprestimos_file_data = None
            
            for part in parts:
                if b'Content-Disposition: form-data' in part:
                    if b'name="taxas_file"' in part:
                        # Extrair arquivo de taxas
                        header_end = part.find(b'\r\n\r\n')
                        if header_end != -1:
                            taxas_file_data = part[header_end + 4:-2]  # Remove \r\n do final
                    elif b'name="emprestimos_file"' in part:
                        # Extrair arquivo de empréstimos
                        header_end = part.find(b'\r\n\r\n')
                        if header_end != -1:
                            emprestimos_file_data = part[header_end + 4:-2]  # Remove \r\n do final
                    elif b'name=' in part:
                        # Extrair campos de formulário
                        lines = part.split(b'\r\n')
                        for line in lines:
                            if line.startswith(b'Content-Disposition: form-data; name='):
                                name_start = line.find(b'name="') + 6
                                name_end = line.find(b'"', name_start)
                                name = line[name_start:name_end].decode('utf-8')
                                
                                # Encontrar o valor na próxima linha
                                value_line = lines[lines.index(line) + 2] if lines.index(line) + 2 < len(lines) else b''
                                form_data[name] = value_line.decode('utf-8').strip()
            
            # Extrair parâmetros
            taxa_juros = float(form_data.get('taxa_juros', 6.5)) / 100.0
            tabua_validos = form_data.get('tabua_validos', 'AT-83')
            tabua_invalidos = form_data.get('tabua_invalidos', 'AT-83')
            
            if not taxas_file_data or not emprestimos_file_data:
                raise ValueError("Arquivos não encontrados no upload")
            
            # Processar arquivo de taxas de risco
            df_taxas = pd.read_excel(io.BytesIO(taxas_file_data), engine='openpyxl')
            
            # Mapear colunas do arquivo de taxas
            colunas_esperadas_taxas = ['idade', 'sexo', 'situacao', 'parcela', 'taxa_risco_mensal']
            colunas_originais_taxas = list(df_taxas.columns)
            
            if not all(col in df_taxas.columns for col in colunas_esperadas_taxas):
                # Mapear colunas
                mapeamento_taxas = {
                    'Idade': 'idade',
                    'Sexo': 'sexo',
                    'Parcelas Restantes': 'parcela',
                    'Tipo Tábua': 'situacao',
                    'Taxa Risco Mensal (%)': 'taxa_risco_mensal'
                }
                
                colunas_mapeadas_taxas = {}
                for col_original, col_nova in mapeamento_taxas.items():
                    if col_original in df_taxas.columns:
                        colunas_mapeadas_taxas[col_original] = col_nova
                
                if colunas_mapeadas_taxas:
                    df_taxas = df_taxas.rename(columns=colunas_mapeadas_taxas)
                    df_taxas['situacao'] = df_taxas['situacao'].map({'Válido': 'valido', 'Inválido': 'invalido'})
                    df_taxas['sexo'] = df_taxas['sexo'].map({'Masculino': 'M', 'Feminino': 'F'})
                    df_taxas['parcela'] = df_taxas['parcela'].astype(int)
                    df_taxas['taxa_risco_mensal'] = df_taxas['taxa_risco_mensal'].astype(float) / 100
                else:
                    raise ValueError("Não foi possível mapear as colunas do arquivo de taxas")
            
            # Processar arquivo de empréstimos
            df_emprestimos = pd.read_excel(io.BytesIO(emprestimos_file_data), engine='openpyxl')
            
            # Mapear colunas do arquivo de empréstimos
            colunas_esperadas_emprestimos = ['saldo_adimplente', 'prazo_restante', 'idade', 'sexo']
            colunas_originais_emprestimos = list(df_emprestimos.columns)
            
            if not all(col in df_emprestimos.columns for col in colunas_esperadas_emprestimos):
                # Mapear colunas
                mapeamento_emprestimos = {
                    'Saldo Adimplente': 'saldo_adimplente',
                    'Prazo Restante': 'prazo_restante',
                    'Idade': 'idade',
                    'Sexo': 'sexo'
                }
                
                colunas_mapeadas_emprestimos = {}
                for col_original, col_nova in mapeamento_emprestimos.items():
                    if col_original in df_emprestimos.columns:
                        colunas_mapeadas_emprestimos[col_original] = col_nova
                
                if colunas_mapeadas_emprestimos:
                    df_emprestimos = df_emprestimos.rename(columns=colunas_mapeadas_emprestimos)
                    df_emprestimos['sexo'] = df_emprestimos['sexo'].map({'Masculino': 'M', 'Feminino': 'F'})
                    df_emprestimos['saldo_adimplente'] = df_emprestimos['saldo_adimplente'].astype(float)
                    df_emprestimos['prazo_restante'] = df_emprestimos['prazo_restante'].astype(int)
                    df_emprestimos['idade'] = df_emprestimos['idade'].astype(int)
                else:
                    raise ValueError("Não foi possível mapear as colunas do arquivo de empréstimos")
            
            # OTIMIZAÇÃO: Usar cache para tábuas de mortalidade
            print(f"Carregando tábuas de mortalidade...")
            tabua_obj_validos = obter_tabua_cached(taxa_juros, tabua_validos)
            tabua_obj_invalidos = obter_tabua_cached(taxa_juros, tabua_invalidos)
            print(f"Tábuas carregadas com sucesso!")
            
            # Calcular reservas matemáticas para cada empréstimo
            resultados = []
            
            for index, row in df_emprestimos.iterrows():
                try:
                    saldo_devedor = float(row['saldo_adimplente'])
                    parcelas_restantes = int(row['prazo_restante'])
                    idade = int(row['idade'])
                    sexo = str(row['sexo'])
                    
                    # Determinar situação e tábua baseado no sexo
                    situacao = 'valido'  # Assumir válido por padrão
                    tabua = tabua_validos if situacao == 'valido' else tabua_invalidos
                    
                    # Usar instância já criada (OTIMIZAÇÃO)
                    tabua_obj = tabua_obj_validos if situacao == 'valido' else tabua_obj_invalidos
                    
                    # Calcular VABF e VACF usando o método otimizado
                    resultado_otimizado = self.calcular_vabf_vacf_otimizado(
                        tabua_obj, saldo_devedor, parcelas_restantes, idade, sexo, 
                        situacao, df_taxas, taxa_juros, tabua
                    )
                    
                    if resultado_otimizado['sucesso']:
                        resultados.append({
                            'saldo_adimplente': saldo_devedor,
                            'prazo_restante': parcelas_restantes,
                            'idade': idade,
                            'sexo': sexo,
                            'vabf': resultado_otimizado['vabf_otimizado'],
                            'vacf': resultado_otimizado['vacf_otimizado'],
                            'reserva_matematica': resultado_otimizado['reserva_otimizada']
                        })
                    else:
                        # Em caso de erro, adicionar valores zero
                        resultados.append({
                            'saldo_adimplente': saldo_devedor,
                            'prazo_restante': parcelas_restantes,
                            'idade': idade,
                            'sexo': sexo,
                            'vabf': 0,
                            'vacf': 0,
                            'reserva_matematica': 0,
                            'erro': resultado_otimizado.get('erro', 'Erro desconhecido')
                        })
                        
                except Exception as e:
                    # Em caso de erro, adicionar valores zero
                    resultados.append({
                        'saldo_adimplente': float(row.get('saldo_adimplente', 0)),
                        'prazo_restante': int(row.get('prazo_restante', 0)),
                        'idade': int(row.get('idade', 0)),
                        'sexo': str(row.get('sexo', '')),
                        'vabf': 0,
                        'vacf': 0,
                        'reserva_matematica': 0,
                        'erro': str(e)
                    })
            
            # Preparar resposta
            response = {
                "success": True,
                "resultados": resultados,
                "total_emprestimos": len(resultados),
                "vabf_total": sum(r['vabf'] for r in resultados),
                "vacf_total": sum(r['vacf'] for r in resultados),
                "reserva_total": sum(r['vabf'] for r in resultados) - sum(r['vacf'] for r in resultados)
            }
            
            self.send_response(200)
            self.send_header('Content-type', 'application/json; charset=utf-8')
            self.end_headers()
            self.wfile.write(json.dumps(response, ensure_ascii=False).encode('utf-8'))
            
        except Exception as e:
            # Log do erro para debug
            error_response = {
                "success": False, 
                "error": f"Erro interno do servidor: {str(e)}",
                "error_type": type(e).__name__
            }
            
            self.send_response(500)
            self.send_header('Content-type', 'application/json; charset=utf-8')
            self.end_headers()
            self.wfile.write(json.dumps(error_response, ensure_ascii=False).encode('utf-8'))

    def handle_preview_planilha(self):
        """Endpoint para preview das primeiras 10 linhas de uma planilha"""
        try:
            import pandas as pd
            import io
            
            content_length = int(self.headers.get('Content-Length', 0))
            if content_length == 0:
                raise ValueError("Content-Length is 0")
            
            post_data = self.rfile.read(content_length)
            if not post_data:
                raise ValueError("No data received")
            
            content_type = self.headers.get('Content-Type', '')
            if 'boundary=' not in content_type:
                raise ValueError("Boundary não encontrado no Content-Type")
            
            boundary = content_type.split('boundary=')[1]
            parts = post_data.split(f'--{boundary}'.encode())
            
            file_data = None
            file_type = None
            
            for part in parts:
                if b'Content-Disposition: form-data' in part:
                    if b'name="file"' in part:
                        header_end = part.find(b'\r\n\r\n')
                        if header_end != -1:
                            file_data = part[header_end + 4:-2]
                    elif b'name="type"' in part:
                        lines = part.split(b'\r\n')
                        for line in lines:
                            if line.startswith(b'Content-Disposition: form-data; name='):
                                value_line = lines[lines.index(line) + 2] if lines.index(line) + 2 < len(lines) else b''
                                file_type = value_line.decode('utf-8').strip()
            
            if not file_data or not file_type:
                raise ValueError("Arquivo ou tipo não encontrado no upload")
            
            # Ler arquivo Excel
            df = pd.read_excel(io.BytesIO(file_data), engine='openpyxl')
            
            # Pegar apenas as primeiras 10 linhas
            df_preview = df.head(10)
            
            # Converter para formato JSON
            colunas = df_preview.columns.tolist()
            dados = df_preview.values.tolist()
            
            # Formatar dados para exibição
            dados_formatados = []
            for linha in dados:
                linha_formatada = []
                for i, celula in enumerate(linha):
                    if pd.isna(celula):
                        linha_formatada.append('')
                    elif isinstance(celula, (int, float)):
                        if file_type == 'taxas' and i == 4:  # Taxa Risco Mensal
                            linha_formatada.append(f"{celula:.6f}%")
                        elif file_type == 'emprestimos' and i == 0:  # Saldo Adimplente
                            linha_formatada.append(f"R$ {celula:,.2f}")
                        else:
                            linha_formatada.append(str(celula))
                    else:
                        linha_formatada.append(str(celula))
                dados_formatados.append(linha_formatada)
            
            response = {
                "success": True,
                "colunas": colunas,
                "dados": dados_formatados,
                "total_linhas": len(df),
                "tipo": file_type
            }
            
            self.send_response(200)
            self.send_header('Content-type', 'application/json; charset=utf-8')
            self.end_headers()
            self.wfile.write(json.dumps(response, ensure_ascii=False).encode('utf-8'))
            
        except Exception as e:
            error_response = {
                "success": False, 
                "error": f"Erro ao processar preview: {str(e)}",
                "error_type": type(e).__name__
            }
            
            self.send_response(500)
            self.send_header('Content-type', 'application/json; charset=utf-8')
            self.end_headers()
            self.wfile.write(json.dumps(error_response, ensure_ascii=False).encode('utf-8'))

    def handle_obter_tabua_completa(self):
        """Endpoint para obter uma tábua completa de mortalidade."""
        try:
            # Ler dados do POST
            content_length = int(self.headers.get('Content-Length', 0))
            if content_length == 0:
                raise ValueError("Content-Length is 0")
            
            post_data = self.rfile.read(content_length)
            if not post_data:
                raise ValueError("No data received")
            
            data = json.loads(post_data.decode('utf-8'))
            tabua_nome = data.get('tabua')
            
            if not tabua_nome:
                raise ValueError("Nome da tábua não fornecido")
            
            # Carregar a tábua completa
            tabua_data = self.carregar_tabua_completa(tabua_nome)
            
            if tabua_data is None:
                raise ValueError(f"Tábua '{tabua_nome}' não encontrada")
            
            # Enviar resposta
            response_data = {
                "success": True,
                "tabua_data": tabua_data
            }
            
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps(response_data).encode('utf-8'))
            
        except Exception as e:
            error_response = {
                "success": False,
                "error": str(e)
            }
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps(error_response).encode('utf-8'))

    def carregar_tabua_completa(self, tabua_nome):
        """Carrega uma tábua completa de mortalidade do CSV."""
        try:
            # Ler o arquivo CSV
            with open('tabuas_mortalidade_completas.csv', 'r', encoding='utf-8') as file:
                reader = csv.DictReader(file)
                
                tabua_data = {}
                for row in reader:
                    if row['Tábua'] == tabua_nome:
                        idade = int(row['Idade'])
                        tabua_data[idade] = {
                            'qx_masc': float(row['qx_masc']),
                            'qx_fem': float(row['qx_fem'])
                        }
                
                return tabua_data if tabua_data else None
                
        except Exception as e:
            print(f"Erro ao carregar tábua {tabua_nome}: {e}")
            return None

    def handle_obter_qx(self):
        """Endpoint para obter probabilidade de morte de uma tábua específica."""
        try:
            # Ler dados do POST
            content_length = int(self.headers.get('Content-Length', 0))
            if content_length == 0:
                raise ValueError("Content-Length is 0")
            
            post_data = self.rfile.read(content_length)
            if not post_data:
                raise ValueError("No data received")
            
            data = json.loads(post_data.decode('utf-8'))
            
            # Extrair parâmetros
            tabua = data.get('tabua', '')
            idade = data.get('idade', 0)
            sexo = data.get('sexo', 'M')
            
            # Validação básica
            if not tabua:
                raise ValueError("Tábua não especificada")
            
            if not (0 <= idade <= 110):
                raise ValueError("Idade deve estar entre 0 e 110 anos")
            
            if sexo not in ['M', 'F']:
                raise ValueError("Sexo deve ser 'M' ou 'F'")
            
            # Criar instância da tábua
            tabua_obj = TabuladeComutacao(0.01, tabua)  # Taxa de juros não importa para obter qx
            
            # Verificar se a tábua existe
            if tabua not in tabua_obj.tabuas_disponiveis:
                raise KeyError(f"Tábua '{tabua}' não encontrada")
            
            # Carregar dados da tábua
            tabua_obj.dados = tabua_obj.tabuas_disponiveis[tabua]
            
            # Obter probabilidade de morte
            qx = tabua_obj.obter_qx(idade, sexo)
            
            response = {
                'success': True,
                'qx': qx
            }
            
            self.send_response(200)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(response, ensure_ascii=False).encode('utf-8'))
            
        except Exception as e:
            error_response = {"success": False, "error": str(e)}
            self.send_response(500)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(error_response, ensure_ascii=False).encode('utf-8'))

    def handle_calcular_reserva_matematica_individual(self):
        """Calcula a reserva matemática individual com upload de planilha XLSX."""
        try:
            import io
            import pandas as pd
            import tempfile
            import os
            import urllib.parse
            
            # Iniciando cálculo de reserva matemática
            
            # Ler dados do POST de forma mais simples
            content_length = int(self.headers.get('Content-Length', 0))
            if content_length == 0:
                raise ValueError("Content-Length is 0")
            
            post_data = self.rfile.read(content_length)
            if not post_data:
                raise ValueError("No data received")
            
            # Parse simples dos dados (assumindo que é form-data)
            # Para simplificar, vamos usar uma abordagem diferente
            # Processando dados do POST e formulário
            
            # Extrair boundary do Content-Type
            content_type = self.headers.get('Content-Type', '')
            if 'boundary=' not in content_type:
                raise ValueError("Boundary não encontrado no Content-Type")
            
            boundary = content_type.split('boundary=')[1].encode()
            
            # Parse manual dos dados multipart
            parts = post_data.split(b'--' + boundary)
            form_data = {}
            file_data = None
            
            for part in parts:
                if b'Content-Disposition' in part:
                    # Extrair nome do campo
                    if b'name="' in part:
                        name_start = part.find(b'name="') + 6
                        name_end = part.find(b'"', name_start)
                        field_name = part[name_start:name_end].decode('utf-8')
                        
                        # Verificar se é arquivo
                        if b'filename="' in part:
                            # Extrair dados do arquivo
                            data_start = part.find(b'\r\n\r\n') + 4
                            data_end = part.rfind(b'\r\n')
                            file_data = part[data_start:data_end]
                        else:
                            # Extrair valor do campo
                            data_start = part.find(b'\r\n\r\n') + 4
                            data_end = part.rfind(b'\r\n')
                            if data_end > data_start:
                                field_value = part[data_start:data_end].decode('utf-8')
                                form_data[field_name] = field_value
            
            # Extrair dados do formulário
            saldo_devedor = float(form_data.get('saldo_devedor', 100000.0))
            parcelas_restantes = int(form_data.get('parcelas_restantes', 12))
            idade = int(form_data.get('idade', 30))
            sexo = form_data.get('sexo', 'M')
            situacao = form_data.get('situacao', 'valido')
            tabua = form_data.get('tabua', 'AT-83')
            taxa_juros = float(form_data.get('taxa_juros', 6.5)) / 100.0  # Converter de % para decimal
            
            # Dados extraídos do formulário
            
            # Processar arquivo XLSX do upload
            if not file_data:
                raise ValueError("Arquivo não encontrado no upload")
            
            # Arquivo recebido
            
            # Ler arquivo XLSX do upload
            try:
                df = pd.read_excel(io.BytesIO(file_data), engine='openpyxl')
            except Exception as e:
                # Erro ao ler arquivo
                raise ValueError(f"Erro ao ler arquivo XLSX: {str(e)}")
            
            # Verificar se as colunas já estão no formato correto
            colunas_esperadas = ['idade', 'sexo', 'situacao', 'parcela', 'taxa_risco_mensal']
            colunas_originais = list(df.columns)
            
            # Verificando colunas do arquivo
            
            # Se as colunas já estão corretas, não precisa mapear
            if not all(col in df.columns for col in colunas_esperadas):
                # Tentar mapear colunas do arquivo para o formato esperado
                mapeamento_colunas = {
                    'Idade': 'idade',
                    'Sexo': 'sexo', 
                    'Parcelas Restantes': 'parcela',
                    'Tipo Tábua': 'situacao',
                    'Taxa Risco Mensal (%)': 'taxa_risco_mensal'
                }
                
                colunas_mapeadas = {}
                
                for col_original, col_nova in mapeamento_colunas.items():
                    if col_original in df.columns:
                        colunas_mapeadas[col_original] = col_nova
                        # Mapeando coluna
                
                if colunas_mapeadas:
                    # Renomear colunas
                    df = df.rename(columns=colunas_mapeadas)
                    
                    # Converter tipos de dados
                    df['idade'] = df['idade'].astype(int)
                    df['sexo'] = df['sexo'].map({'Masculino': 'M', 'Feminino': 'F'})
                    df['situacao'] = df['situacao'].map({'Válido': 'valido', 'Inválido': 'invalido'})
                    df['parcela'] = df['parcela'].astype(int)
                    df['taxa_risco_mensal'] = df['taxa_risco_mensal'].astype(float) / 100  # Converter de % para decimal
                else:
                    raise ValueError("Não foi possível mapear as colunas do arquivo")
            
            # Arquivo processado com sucesso
            
            # Criar instância da tábua de comutação
            tabua_obj = TabuladeComutacao(taxa_juros, tabua)
            # Tábua selecionada e processada
            
            # Calcular VABF e VACF
            import time
            
            # Calcular VABF e VACF
            resultado = self.calcular_vabf_vacf_individual(
                tabua_obj, saldo_devedor, parcelas_restantes, idade, sexo, 
                situacao, df, taxa_juros, tabua
            )
            
            # Calcular também a versão otimizada para comparação
            resultado_otimizado = self.calcular_vabf_vacf_otimizado(
                tabua_obj, saldo_devedor, parcelas_restantes,
                idade, sexo, situacao, df, taxa_juros, tabua
            )
            
            # Calcular diferenças para comparação
            if resultado_otimizado['sucesso']:
                diferenca_vabf = abs(resultado['vabf'] - resultado_otimizado['vabf_otimizado'])
                diferenca_vacf = abs(resultado['vacf'] - resultado_otimizado['vacf_otimizado'])
                diferenca_reserva = abs(resultado['reserva_matematica'] - resultado_otimizado['reserva_otimizada'])
                
                # Calcular percentuais de diferença
                percentual_vabf = (diferenca_vabf / resultado['vabf'] * 100) if resultado['vabf'] != 0 else 0
                percentual_vacf = (diferenca_vacf / resultado['vacf'] * 100) if resultado['vacf'] != 0 else 0
                percentual_reserva = (diferenca_reserva / abs(resultado['reserva_matematica']) * 100) if resultado['reserva_matematica'] != 0 else 0
                
                comparacao = {
                    'diferenca_vabf': diferenca_vabf,
                    'diferenca_vacf': diferenca_vacf,
                    'diferenca_reserva': diferenca_reserva,
                    'percentual_diferenca_vabf': percentual_vabf,
                    'percentual_diferenca_vacf': percentual_vacf,
                    'percentual_diferenca_reserva': percentual_reserva
                }
            else:
                comparacao = None
            
            response = {
                "success": True,
                "vabf": resultado['vabf'],
                "vacf": resultado['vacf'],
                "reserva_matematica": resultado['reserva_matematica'],
                "detalhes": resultado['detalhes'],
                "tabua": tabua,
                "vabf_otimizado": resultado_otimizado.get('vabf_otimizado'),
                "vacf_otimizado": resultado_otimizado.get('vacf_otimizado'),
                "reserva_otimizada": resultado_otimizado.get('reserva_otimizada'),
                "comparacao": comparacao
            }
            
            # Enviar resposta
            self.send_response(200)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(response, ensure_ascii=False).encode('utf-8'))
            
        except Exception as e:
            # Log do erro para debug
            # Erro no cálculo da reserva matemática
            
            error_response = {
                "success": False, 
                "error": f"Erro interno do servidor: {str(e)}",
                "error_type": type(e).__name__
            }
            self.send_response(500)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(error_response, ensure_ascii=False).encode('utf-8'))

    def calcular_vabf_vacf_individual(self, tabua_obj, saldo_devedor, parcelas_restantes, 
                                    idade, sexo, situacao, df_taxas, taxa_juros, tabua):
        """
        Calcula VABF (Valor Atual dos Benefícios Futuros) e VACF (Valor Atual dos Custos Futuros)
        para reserva matemática individual do seguro prestamista.
        """
        import math
        
        # Fator de desconto mensal será calculado depois da taxa_mensal
        vabf = 0.0
        vacf = 0.0
        
        # Calcular taxa mensal corretamente
        taxa_mensal = (1 + taxa_juros)**(1/12) - 1
        
        # Fator de desconto mensal
        v = 1 / (1 + taxa_mensal)
        
        # Calcular parcela mensal do financiamento Price usando taxa mensal
        if taxa_mensal > 0:
            parcela_mensal = saldo_devedor * (taxa_mensal * (1 + taxa_mensal)**parcelas_restantes) / ((1 + taxa_mensal)**parcelas_restantes - 1)
        else:
            parcela_mensal = saldo_devedor / parcelas_restantes
        
        detalhes = {
            'parcela_mensal': parcela_mensal,
            'saldo_inicial': saldo_devedor,
            'parcelas_restantes': parcelas_restantes,
            'idade_inicial': idade,
            'sexo': sexo,
            'situacao': situacao,
            'taxa_juros_mensal': taxa_mensal,
            'taxa_juros_anual': taxa_juros,  # Já está em decimal (ex: 0.065)
            'tabua_mortalidade': tabua,
            'calculo_mensal': []
        }
        
        # Calcular VABF e VACF mês a mês
        saldo_atual = saldo_devedor
        idade_inicial = idade  # Armazenar a idade inicial para cálculo de taxa de risco
        idade_atual = idade
        
        for mes in range(1, parcelas_restantes + 1):
            # Calcular saldo devedor no mês atual (B_{t-1})
            if taxa_mensal > 0:
                juros_mes = saldo_atual * taxa_mensal
                amortizacao_mes = parcela_mensal - juros_mes
            else:
                juros_mes = 0
                amortizacao_mes = parcela_mensal
            
            saldo_proximo = max(0, saldo_atual - amortizacao_mes)
            
            # Verificar se a idade mudou (a cada 12 meses)
            if mes > 1 and (mes - 1) % 12 == 0:
                idade_atual += 1
            
            # Obter taxa de risco do CSV
            taxa_risco = self.obter_taxa_risco_csv(df_taxas, idade_inicial, sexo, situacao, mes, parcelas_restantes)
            
            # Obter qx anual da tábua para a idade atual
            qx_anual = tabua_obj.obter_qx(idade_atual, sexo)
            
            # Fórmulas corretas conforme explicado pelo usuário:
            # P_mensal = (1 - q_anual)^(1/12)
            # q_mensal = 1 - P_mensal
            p_mensal = (1 - qx_anual) ** (1/12)
            qx_mensal = 1 - p_mensal
            
            # VABF: Valor presente do benefício usando fórmula correta
            # VABF = Σ(t=1 to m) B_{t-1} × _{t-1}P_x × q_{x+t-1} × v^t
            
            # B_{t-1}: Saldo devedor no início do período t (saldo_atual)
            saldo_devedor_t_menos_1 = saldo_atual
            
            # _{t-1}P_x: Probabilidade de sobrevivência até t-1
            # Calcular considerando mudanças de idade a cada 12 meses
            if mes == 1:
                prob_sobrevivencia_t_menos_1 = 1.0
            else:
                # Calcular _{t-1}P_x = produto das probabilidades mensais até t-1
                prob_sobrevivencia_t_menos_1 = 1.0
                
                for m in range(1, mes):
                    # Determinar a idade no mês m
                    idade_mes_m = idade + ((m - 1) // 12)
                    
                    # Obter qx anual para essa idade
                    qx_anual_mes_m = tabua_obj.obter_qx(idade_mes_m, sexo)
                    
                    # Converter para mensal
                    p_mensal_mes_m = (1 - qx_anual_mes_m) ** (1/12)
                    
                    # Multiplicar pela probabilidade de sobrevivência
                    prob_sobrevivencia_t_menos_1 *= p_mensal_mes_m
            
            # q_{x+t-1}: Probabilidade de morte no período t (qx_mensal da idade atual)
            qx_t_menos_1 = qx_mensal
            
            # VABF = B_{t-1} × _{t-1}P_x × q_{x+t-1} × v^t
            vabf_mes = saldo_devedor_t_menos_1 * prob_sobrevivencia_t_menos_1 * qx_t_menos_1 * (v ** mes)
            vabf += vabf_mes
            
            # Debug removido para limpeza do terminal
            
            # VACF: Valor presente do prêmio (taxa de risco * saldo devedor) se sobrevivência
            premio_mes = taxa_risco * saldo_atual
            
            # Calcular probabilidade de sobrevivência até o final do mês atual
            # P_{t} = P_{t-1} × p_mensal_atual
            prob_sobrevivencia_fim_mes = prob_sobrevivencia_t_menos_1 * p_mensal
            
            # tPx = multiplicação acumulada de todos os P_{t-1} até o mês atual
            # Para o mês 1: tPx = P_{0} = 1.0
            # Para o mês 2: tPx = P_{0} × P_{1} = 1.0 × p_mensal_mes1
            # Para o mês 3: tPx = P_{0} × P_{1} × P_{2} = 1.0 × p_mensal_mes1 × p_mensal_mes2
            tpx = prob_sobrevivencia_t_menos_1 * p_mensal
            
            vacf_mes = (v ** mes) * prob_sobrevivencia_fim_mes * premio_mes
            vacf += vacf_mes
            
            # Armazenar detalhes do mês
            detalhes['calculo_mensal'].append({
                'mes': mes,
                'idade': idade_atual,
                'saldo_devedor': saldo_atual,
                'saldo_devedor_t_menos_1': saldo_devedor_t_menos_1,
                'amortizacao': amortizacao_mes,
                'juros': juros_mes,
                'taxa_risco': taxa_risco,
                'prob_sobrevivencia': prob_sobrevivencia_fim_mes,
                'prob_sobrevivencia_t_menos_1': prob_sobrevivencia_t_menos_1,
                'tpx': tpx,  # Adicionado tPx
                'qx': qx_mensal,
                'qx_anual': qx_anual,
                'p_mensal': p_mensal,
                'qx_t_menos_1': qx_t_menos_1,
                'beneficio': saldo_devedor_t_menos_1,
                'premio': premio_mes,
                'fator_desconto': v ** mes,
                'vabf_mes': vabf_mes,
                'vacf_mes': vacf_mes
            })
            
            # Atualizar para próximo mês
            saldo_atual = saldo_proximo
    
        # Calcular reserva matemática
        reserva_matematica = vabf - vacf
        
        return {
            'vabf': vabf,
            'vacf': vacf,
            'reserva_matematica': reserva_matematica,
            'detalhes': detalhes
        }

    def obter_taxa_risco_csv(self, df, idade_inicial, sexo, situacao, mes, parcelas_restantes):
        """
        Obtém a taxa de risco do CSV baseada na lógica correta:
        - Meses 1-12: idade_inicial + parcelas_restantes
        - Meses 13-24: idade_inicial + 1 + (parcelas_restantes - 12)
        - Meses 25-36: idade_inicial + 2 + (parcelas_restantes - 24)
        E assim por diante...
        """
        # Calcular idade e parcela para buscar na planilha
        anos_transcorridos = (mes - 1) // 12
        idade_para_taxa = idade_inicial + anos_transcorridos
        parcela_para_taxa = parcelas_restantes - (anos_transcorridos * 12)
        
        # Debug removido para limpeza do terminal
        
        # Filtrar dados do CSV
        filtro = (df['idade'] == idade_para_taxa) & (df['sexo'] == sexo) & (df['situacao'] == situacao) & (df['parcela'] == parcela_para_taxa)
        dados_filtrados = df[filtro]
        
        if len(dados_filtrados) == 0:
            # Se não encontrar exato, buscar o mais próximo
            for idade_diff in range(1, 6):  # Buscar até 5 anos de diferença
                for idade_tentativa in [idade_para_taxa - idade_diff, idade_para_taxa + idade_diff]:
                    if idade_tentativa >= 0:
                        filtro_alt = (df['idade'] == idade_tentativa) & (df['sexo'] == sexo) & (df['situacao'] == situacao) & (df['parcela'] == parcela_para_taxa)
                        dados_alt = df[filtro_alt]
                        if len(dados_alt) > 0:
                            return float(dados_alt.iloc[0]['taxa_risco_mensal'])
            
            # Se ainda não encontrar, usar taxa padrão baseada na idade
            if idade_para_taxa < 30:
                return 0.0001  # 0.01% ao mês
            elif idade_para_taxa < 50:
                return 0.0005  # 0.05% ao mês
            elif idade_para_taxa < 70:
                return 0.001   # 0.1% ao mês
            else:
                return 0.002   # 0.2% ao mês
        
        return float(dados_filtrados.iloc[0]['taxa_risco_mensal'])

    def calcular_vabf_vacf_otimizado(self, tabua_obj, saldo_devedor, parcelas_restantes,
                                     idade, sexo, situacao, df, taxa_juros, tabua):
        """
        Versão ULTRA OTIMIZADA usando técnicas de comutação e vetorização máxima.
        Elimina loops desnecessários e usa lookup tables para máxima performance.
        """
        try:
            import numpy as np
            import pandas as pd
            
            # ===== PRÉ-COMPUTAÇÕES BÁSICAS =====
            taxa_mensal = (1 + taxa_juros)**(1/12) - 1
            v = 1 / (1 + taxa_mensal)
            
            # Calcular parcela mensal
            if taxa_mensal > 0:
                parcela_mensal = saldo_devedor * (taxa_mensal * (1 + taxa_mensal)**parcelas_restantes) / ((1 + taxa_mensal)**parcelas_restantes - 1)
            else:
                parcela_mensal = saldo_devedor / parcelas_restantes
            
            # ===== CÁLCULO CORRETO DO SALDO DEVEDOR (PRICE) =====
            # Calcular saldo devedor mês a mês usando a mesma lógica do método iterativo
            saldos_devedor = np.zeros(parcelas_restantes)
            saldo_atual = saldo_devedor
            
            for i in range(parcelas_restantes):
                saldos_devedor[i] = saldo_atual
                if taxa_mensal > 0:
                    juros_mes = saldo_atual * taxa_mensal
                    amortizacao_mes = parcela_mensal - juros_mes
                else:
                    juros_mes = 0
                    amortizacao_mes = parcela_mensal
                saldo_atual = max(0, saldo_atual - amortizacao_mes)
            
            # ===== PRÉ-COMPUTAÇÃO DAS IDADES E PROBABILIDADES =====
            meses = np.arange(1, parcelas_restantes + 1)
            idades_mes = idade + ((meses - 1) // 12)
            
            # Obter qx anual para todas as idades de uma vez (ULTRA VETORIZADO)
            # Usar vectorize para eliminar o laço
            obter_qx_vectorized = np.vectorize(lambda idade_m: tabua_obj.obter_qx(int(idade_m), sexo))
            qx_anuais = obter_qx_vectorized(idades_mes)
            
            # Converter para probabilidades mensais vetorizado
            p_mensais = (1 - qx_anuais) ** (1/12)
            qx_mensais = 1 - p_mensais
            
            # ===== CÁLCULO ULTRA VETORIZADO DE _{t-1}P_x =====
            # Usar cumprod para calcular produto acumulado de forma vetorizada
            prob_sobrevivencia_acumulada = np.ones(parcelas_restantes)
            if parcelas_restantes > 1:
                prob_sobrevivencia_acumulada[1:] = np.cumprod(p_mensais[:-1])
            
            # ===== CÁLCULO ULTRA VETORIZADO DAS TAXAS DE RISCO =====
            anos_transcorridos = (meses - 1) // 12
            idades_para_taxa = idade + anos_transcorridos
            parcelas_para_taxa = parcelas_restantes - (anos_transcorridos * 12)
            
            # ULTRA OTIMIZAÇÃO: Usar merge do pandas para lookup vetorizado
            if df is not None and not df.empty:
                # Criar DataFrame com as chaves de busca
                df_busca = pd.DataFrame({
                    'idade': idades_para_taxa,
                    'sexo': [sexo] * parcelas_restantes,
                    'situacao': [situacao] * parcelas_restantes,
                    'parcela': parcelas_para_taxa
                })
                
                # Merge vetorizado para obter taxas de risco
                df_merged = df_busca.merge(df, on=['idade', 'sexo', 'situacao', 'parcela'], how='left')
                taxas_risco = df_merged['taxa_risco_mensal'].fillna(0.001).values
            else:
                # Taxa padrão vetorizada baseada na idade
                taxas_risco = np.where(idades_para_taxa < 30, 0.0001,
                                     np.where(idades_para_taxa < 50, 0.0005,
                                             np.where(idades_para_taxa < 70, 0.001, 0.002)))
            
            # ===== CÁLCULOS FINAIS CORRIGIDOS =====
            # Pré-computar fatores de desconto
            fatores_desconto = v ** meses
            
            # Probabilidade de morte no mês t (como no método iterativo)
            prob_morte = prob_sobrevivencia_acumulada * qx_mensais
            
            # VABF = Σ B_{t-1} × _{t-1}P_x × q_{x+t-1} × v^t
            vabf_otimizado = np.sum(saldos_devedor * prob_morte * fatores_desconto)
            
            # VACF = Σ premio_mes × prob_sobrevivencia_fim_mes × v^t
            # onde premio_mes = taxa_risco × saldo_devedor
            premios_mensais = taxas_risco * saldos_devedor
            prob_sobrevivencia_fim_mes = prob_sobrevivencia_acumulada * p_mensais
            vacf_otimizado = np.sum(premios_mensais * prob_sobrevivencia_fim_mes * fatores_desconto)
            
            # Calcular reserva matemática
            reserva_otimizada = vabf_otimizado - vacf_otimizado
            
            return {
                'vabf_otimizado': vabf_otimizado,
                'vacf_otimizado': vacf_otimizado,
                'reserva_otimizada': reserva_otimizada,
                'sucesso': True
            }
            
        except Exception as e:
            return {
                'vabf_otimizado': None,
                'vacf_otimizado': None,
                'reserva_otimizada': None,
                'sucesso': False,
                'erro': str(e)
            }

def obter_ip_rede_local():
    """Obtém o IP da rede local automaticamente."""
    import socket
    try:
        # Conectar a um endereço externo para descobrir o IP local
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip_local = s.getsockname()[0]
        s.close()
        return ip_local
    except:
        return "127.0.0.1"

def encontrar_porta_disponivel(porta_inicial=PORT_INICIAL, max_tentativas=100):
    """
    Encontra uma porta disponível começando da porta inicial
    """
    for porta in range(porta_inicial, porta_inicial + max_tentativas):
        try:
            # Tenta criar um socket para testar se a porta está disponível
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                s.bind(('0.0.0.0', porta))
                return porta
        except OSError:
            # Porta em uso, tenta a próxima
            continue
    
    # Se não encontrou nenhuma porta disponível
    raise RuntimeError(f"Não foi possível encontrar uma porta disponível entre {porta_inicial} e {porta_inicial + max_tentativas - 1}")

if __name__ == '__main__':
    # Obter IP da rede local
    ip_local = obter_ip_rede_local()
    
    # Encontrar uma porta disponível
    PORT = encontrar_porta_disponivel()
    
    # Informar sobre a porta sendo usada
    if PORT != PORT_INICIAL:
        print(f"AVISO: Porta {PORT_INICIAL} esta em uso. Usando porta {PORT}.")
    
    with socketserver.TCPServer(("0.0.0.0", PORT), CalculadoraHandler) as httpd:
        print("=" * 60)
        print("SERVIDOR DE SEGURO PRESTAMISTA INICIADO")
        print("=" * 60)
        print(f"Calculadora de Seguro de Vida - Web")
        print(f"Servidor rodando na porta: {PORT}")
        print()
        print("ACESSO LOCAL:")
        print(f"   • http://localhost:{PORT}")
        print(f"   • http://127.0.0.1:{PORT}")
        print()
        print("ACESSO NA REDE LOCAL:")
        print(f"   • http://{ip_local}:{PORT}")
        print()
        print("Para acessar de outro dispositivo na mesma rede WiFi:")
        print(f"   1. Conecte o dispositivo na mesma rede WiFi")
        print(f"   2. Abra o navegador no dispositivo")
        print(f"   3. Digite: http://{ip_local}:{PORT}")
        print()
        print("IMPORTANTE:")
        print("   • Certifique-se de que o firewall permite conexões na porta", PORT)
        print("   • Ambos os dispositivos devem estar na mesma rede WiFi")
        print("   • Se não funcionar, verifique as configurações do firewall")
        print()
        print("Pressione Ctrl+C para parar o servidor")
        print("=" * 60)
        
        try:
            httpd.serve_forever()
        except KeyboardInterrupt:
            print("\n🛑 Servidor parado.")
